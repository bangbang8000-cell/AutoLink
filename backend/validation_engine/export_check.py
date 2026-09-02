"""F5-2（45-a/45-e）导出数据核对：渲染批次（output/<batch>）产物 vs 实际设计/规划状态

核对维度：
  E001 批次存在性；E002 manifest 缺失提示
  E003 渲染模式 vs 设计模式漂移
  E004 manifest 统计（servers/leaves/spines/…）vs 设计状态漂移
  E005 连接表行数 vs 设计连接数（或非空合理性）
  E006 产物文件命名与设计模式匹配
  E007 设备清单/BOM 条目非空
  E008 批次完整性（manifest.files 逐文件：缺失/漂移/哈希不符）——48-e（F8-5）

collect_batch_stats 读取 output/<batch> 目录（manifest.json + xlsx 产物行数），
供 check_export_batch 与门禁脚本复用。纯标准库 + openpyxl/pandas 读取，只读不改产物。
"""
import hashlib
import os
import re
from typing import Any, Dict, List, Optional

from validation_engine.core import (
    ValidationProblem, SEVERITY_ERROR, SEVERITY_WARNING, SEVERITY_INFO,
    CATEGORY_EXPORT, CATEGORY_RENDER,
)

# 渲染产物文件名前缀（与 backend/engine.py handle_export 一致）
_FILE_PREFIXES = {
    'connections': 'AI智算网络',
    'deviceList': '设备清单',
    'cablingGuide': '布线指导表',
    'bom': 'BOM成本估算',
    'pdfReport': '设计报告',
}


def _read_json_file(path: str) -> Optional[Dict]:
    try:
        with open(path, 'r', encoding='utf-8') as f:
            import json
            return json.load(f)
    except (OSError, ValueError):
        return None


def _xlsx_data_rows(path: str) -> int:
    """统计 xlsx 数据行数（首行为表头；合并单元格按逻辑行计）"""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        count = 0
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue
            if any(v is not None and v != '' for v in row):
                count += 1
        wb.close()
        return count
    except Exception:
        return -1


def collect_batch_stats(batch_dir: str) -> Dict[str, Any]:
    """收集 output/<batch> 批次统计（不抛异常，缺失字段回退默认）。"""
    stats: Dict[str, Any] = {
        'exists': os.path.isdir(batch_dir),
        'files': [],
        'has_manifest': False,
        'has_files_manifest': False,
        'manifest': None,
        'mode': '',
        'output_types': [],
        'connection_rows': -1,
        'device_list_rows': -1,
        'bom_rows': -1,
    }
    if not stats['exists']:
        return stats
    try:
        stats['files'] = sorted(f for f in os.listdir(batch_dir)
                                if os.path.isfile(os.path.join(batch_dir, f)))
    except OSError:
        return stats

    manifest_path = os.path.join(batch_dir, 'manifest.json')
    if os.path.exists(manifest_path):
        manifest = _read_json_file(manifest_path)
        if isinstance(manifest, dict):
            stats['has_manifest'] = True
            stats['manifest'] = manifest
            stats['mode'] = manifest.get('downlink_mode') or ''
            stats['output_types'] = list(manifest.get('output_types') or [])
            # 48-e（F8-5）：逐文件清单存在标记（完整性校验前置）
            stats['has_files_manifest'] = isinstance(manifest.get('files'), list) and len(manifest.get('files') or []) > 0

    # 文件名推断模式（当 manifest 缺失时兜底）：AI智算网络_full模式_xxx.xlsx
    if not stats['mode']:
        for f in stats['files']:
            m = re.search(r'([a-z]+)模式', f)
            if m:
                stats['mode'] = m.group(1)
                break

    for key, prefix in _FILE_PREFIXES.items():
        for f in stats['files']:
            if f.startswith(prefix) and f.lower().endswith('.xlsx'):
                rows = _xlsx_data_rows(os.path.join(batch_dir, f))
                if key == 'connections':
                    stats['connection_rows'] = rows
                elif key == 'deviceList':
                    stats['device_list_rows'] = rows
                elif key == 'bom':
                    stats['bom_rows'] = rows
                break
    return stats


def check_batch_integrity(batch_dir: str) -> List[ValidationProblem]:
    """48-e（F8-5）：批次完整性——manifest.files 逐文件校验（缺失 / 哈希不符 / 大小漂移 / 清单外文件）。

    读 manifest.json 的 files（name/size/sha256），与磁盘实际文件比对：
      - 清单有、磁盘无 → 缺失（ERROR）
      - 清单有、磁盘有但 sha256 不符 → 哈希不符（ERROR）；size 漂移 → WARNING
      - 磁盘有、清单无 → 漂移（WARNING）
    """
    problems: List[ValidationProblem] = []
    manifest = _read_json_file(os.path.join(batch_dir, 'manifest.json'))
    if not isinstance(manifest, dict) or not isinstance(manifest.get('files'), list):
        return problems
    expected = {f['name']: f for f in manifest['files'] if isinstance(f, dict) and f.get('name')}
    if not expected:
        return problems

    actual: Dict[str, int] = {}
    for fname in os.listdir(batch_dir):
        fpath = os.path.join(batch_dir, fname)
        if os.path.isfile(fpath):
            actual[fname] = os.path.getsize(fpath)

    # 缺失
    for name in sorted(set(expected) - set(actual)):
        problems.append(ValidationProblem(
            rule_id='E008',
            severity=SEVERITY_ERROR,
            category=CATEGORY_EXPORT,
            location=f'output/<batch>/{name}',
            message=f'清单文件缺失: {name}',
            suggestion='重新渲染恢复该产物，或核对是否被误删',
            data={'file': name},
        ))

    # 哈希不符 / 大小漂移
    for name in sorted(set(actual) & set(expected)):
        if name == 'manifest.json':
            continue
        exp = expected[name]
        try:
            with open(os.path.join(batch_dir, name), 'rb') as f:
                digest = hashlib.sha256(f.read()).hexdigest()
        except OSError:
            continue
        exp_hash = str(exp.get('sha256') or '')
        if exp_hash and digest != exp_hash:
            problems.append(ValidationProblem(
                rule_id='E008',
                severity=SEVERITY_ERROR,
                category=CATEGORY_EXPORT,
                location=f'output/<batch>/{name}',
                message=f'产物哈希不符（可能被篡改/损坏）: {name}',
                suggestion='重新渲染生成该产物，避免使用不一致的批次',
                data={'file': name, 'expected_sha256': exp_hash[:12], 'actual_sha256': digest[:12]},
            ))
            continue
        exp_size = exp.get('size')
        if isinstance(exp_size, int) and actual[name] != exp_size:
            problems.append(ValidationProblem(
                rule_id='E008',
                severity=SEVERITY_WARNING,
                category=CATEGORY_EXPORT,
                location=f'output/<batch>/{name}',
                message=f'产物大小漂移: {name}（清单 {exp_size} 字节 / 实际 {actual[name]} 字节）',
                suggestion='核对产物是否被部分改写，必要时重新渲染',
                data={'file': name, 'expected_size': exp_size, 'actual_size': actual[name]},
            ))

    # 清单外文件（漂移）
    for name in sorted(set(actual) - set(expected) - {'manifest.json'}):
        problems.append(ValidationProblem(
            rule_id='E008',
            severity=SEVERITY_WARNING,
            category=CATEGORY_EXPORT,
            location=f'output/<batch>/{name}',
            message=f'清单外文件（漂移）: {name}',
            suggestion='清理无关产物或重新渲染保持清单一致',
            data={'file': name},
        ))

    return problems


def check_export_batch(batch_dir: str,
                       design: Optional[Dict] = None) -> List[ValidationProblem]:
    """核对渲染批次产物与实际设计/规划状态一致性，发现漂移即报问题。"""
    problems: List[ValidationProblem] = []
    stats = collect_batch_stats(batch_dir)

    # --- E001 批次存在性 ---
    if not stats['exists']:
        problems.append(ValidationProblem(
            rule_id='E001',
            severity=SEVERITY_ERROR,
            category=CATEGORY_EXPORT,
            location=f'output/<batch>（{batch_dir}）',
            message='渲染批次目录不存在，无法核对产物',
            suggestion='先渲染导出，再执行导出数据核对',
            data={'batch_dir': batch_dir},
        ))
        return problems

    # --- E002 manifest 缺失提示 ---
    if not stats['has_manifest']:
        problems.append(ValidationProblem(
            rule_id='E002',
            severity=SEVERITY_WARNING,
            category=CATEGORY_EXPORT,
            location='output/<batch>/manifest.json',
            message='批次缺少 manifest.json，无法核对版本/配置哈希/统计',
            suggestion='重新渲染生成带 manifest.json 的批次，确保可追溯',
            data={},
        ))

    # --- E003 渲染模式 vs 设计模式 ---
    manifest = stats['manifest']
    design_mode = (design or {}).get('mode', '') if isinstance(design, dict) else ''
    if stats['mode'] and design_mode and stats['mode'] != design_mode:
        problems.append(ValidationProblem(
            rule_id='E003',
            severity=SEVERITY_ERROR,
            category=CATEGORY_EXPORT,
            location='output/<batch> 文件名/manifest.downlink_mode ↔ design.mode',
            message=f'渲染批次模式 {stats["mode"]} 与设计模式 {design_mode} 漂移',
            suggestion='以当前设计重新渲染，删除旧模式批次或标记过期',
            data={'render_mode': stats['mode'], 'design_mode': design_mode},
        ))

    # --- E004 manifest 统计 vs 设计状态 ---
    if manifest:
        mstats = manifest.get('stats')
        if isinstance(mstats, dict) and isinstance(design, dict):
            nd = design.get('network_devices') if isinstance(design.get('network_devices'), dict) else {}
            pairs = (
                ('servers', 'servers', '服务器'),
                ('param_leaves', 'param_leaves', '参数网 Leaf'),
                ('param_spines', 'param_spines', '参数网 Spine'),
                ('param_cores', 'param_cores', '参数网 Core'),
                ('storage_leaves', 'storage_leaves', '存储网 Leaf'),
                ('storage_spines', 'storage_spines', '存储网 Spine'),
                ('biz_access', 'biz_access', '业务接入'),
                ('biz_agg', 'biz_agg', '业务汇聚'),
                ('oob_access', 'oob_access', 'OOB 接入'),
                ('oob_agg', 'oob_agg', 'OOB 汇聚'),
            )
            for mkey, dkey, label in pairs:
                mval = mstats.get(mkey)
                dval = nd.get(dkey)
                if mval is None or dval is None:
                    continue
                try:
                    if int(mval) != int(dval):
                        problems.append(ValidationProblem(
                            rule_id='E004',
                            severity=SEVERITY_ERROR,
                            category=CATEGORY_EXPORT,
                            location=f'manifest.stats.{mkey} ↔ design.network_devices.{dkey}',
                            message=f'渲染批次 {label}数（{mval}）与当前设计（{dval}）漂移',
                            suggestion='设计变更后需重新渲染；或检查是否误用了旧批次',
                            data={'manifest': mval, 'design': dval, 'key': mkey},
                        ))
                except (TypeError, ValueError):
                    continue
        # 产物结果状态：存在 error 产物
        for r in (manifest.get('results') or []):
            if isinstance(r, dict) and r.get('status') == 'error':
                problems.append(ValidationProblem(
                    rule_id='E004',
                    severity=SEVERITY_WARNING,
                    category=CATEGORY_EXPORT,
                    location='manifest.results[]',
                    message=f'渲染产物 {r.get("type", "?")} 生成失败',
                    suggestion='重新生成该产物并核对',
                    data={'type': r.get('type')},
                ))

    # --- E005 连接表行数 ---
    conn_rows = stats['connection_rows']
    design_conns = (design or {}).get('connections', []) if isinstance(design, dict) else []
    if isinstance(design_conns, list) and design_conns:
        if conn_rows >= 0 and len(design_conns) != conn_rows:
            problems.append(ValidationProblem(
                rule_id='E005',
                severity=SEVERITY_ERROR,
                category=CATEGORY_EXPORT,
                location='连接表.xlsx 行数 ↔ design.connections',
                message=f'连接表行数 {conn_rows} 与设计连接数 {len(design_conns)} 不一致',
                suggestion='核对连接表导出是否完整，或设计变更后重新渲染',
                data={'xlsx_rows': conn_rows, 'design_connections': len(design_conns)},
            ))
    elif conn_rows == 0:
        # 无设计连接数据可对时，0 行视为异常（产物应有数据）
        files = [f for f in stats['files'] if f.startswith(_FILE_PREFIXES['connections'])]
        if files:
            problems.append(ValidationProblem(
                rule_id='E005',
                severity=SEVERITY_WARNING,
                category=CATEGORY_EXPORT,
                location='连接表.xlsx',
                message='连接表产物为空（0 行数据）',
                suggestion='检查网络设计是否生成连接后重新渲染',
                data={},
            ))

    # --- E006 文件命名与模式匹配 ---
    if stats['mode']:
        mismatched = [f for f in stats['files']
                      if '模式' in f and stats['mode'] not in f]
        if mismatched:
            problems.append(ValidationProblem(
                rule_id='E006',
                severity=SEVERITY_WARNING,
                category=CATEGORY_EXPORT,
                location='output/<batch> 文件名',
                message=f'存在与模式 {stats["mode"]} 不匹配的产物文件：{mismatched[0]}',
                suggestion='清理过期产物或重新渲染统一命名',
                data={'mode': stats['mode'], 'mismatched': mismatched[:5]},
            ))

    # --- E007 设备清单/BOM 条目非空 ---
    if 'deviceList' in stats['output_types'] or any(f.startswith(_FILE_PREFIXES['deviceList']) for f in stats['files']):
        if stats['device_list_rows'] == 0:
            problems.append(ValidationProblem(
                rule_id='E007',
                severity=SEVERITY_ERROR,
                category=CATEGORY_EXPORT,
                location='设备清单.xlsx',
                message='设备清单产物为空（0 条目）',
                suggestion='检查设备清单导出是否成功',
                data={},
            ))
    if 'bom' in stats['output_types'] or any(f.startswith(_FILE_PREFIXES['bom']) for f in stats['files']):
        if stats['bom_rows'] == 0:
            problems.append(ValidationProblem(
                rule_id='E007',
                severity=SEVERITY_WARNING,
                category=CATEGORY_EXPORT,
                location='BOM.xlsx',
                message='BOM 产物为空（0 条目）',
                suggestion='检查 BOM 导出是否成功',
                data={},
            ))

    # --- E008 批次完整性（48-e F8-5）：manifest.files 逐文件校验（缺失/漂移/哈希不符） ---
    if stats.get('has_files_manifest'):
        problems.extend(check_batch_integrity(batch_dir))

    return problems
