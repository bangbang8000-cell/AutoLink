"""
AutoLink V2.1 - Excel导出与格式化
提供服务器视角、交换机视角的连接表生成和Excel美化
支持机柜编号和U位范围 (V2.1新增)
"""

import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.utils import get_column_letter


def _extract_number(s):
    """从字符串中提取数字用于排序"""
    if isinstance(s, (int, float)):
        return int(s)
    if not isinstance(s, str):
        return 0
    match = re.search(r'\d+', s)
    return int(match.group()) if match else 0


def _parse_speed_gbps(speed_str: str) -> float:
    """将速率字符串（如 '400G'）解析为 Gbps 数值 (V2.9.3-T6)"""
    if not speed_str:
        return 400.0
    s = str(speed_str).strip().upper()
    for unit, factor in (('GB', 1.0), ('G', 1.0), ('TB', 1000.0), ('T', 1000.0)):
        if s.endswith(unit):
            try:
                return float(s[:-len(unit)]) * factor
            except ValueError:
                break
    try:
        return float(s)
    except ValueError:
        return 400.0


def _conv_to_dict(r):
    """ConvergenceResult -> dict (V2.9.3-T6)"""
    return {
        'networkType': r.network_type,
        'downlinkBwGbps': r.downlink_bw_gbps,
        'uplinkBwGbps': r.uplink_bw_gbps,
        'convergenceRatio': r.convergence_ratio,
        'isBlocking': r.is_blocking,
        'targetRatio': r.target_ratio,
        'meetsTarget': r.meets_target,
        'recommendation': r.recommendation,
    }


def _compute_convergence(designer) -> dict:
    """V2.9.3-T6: 计算参数/存储/业务网收敛比(与 engine._estimate_design 一致)"""
    from estimation import calc_convergence_ratio

    convergence = {}
    if designer.param_leaf_count > 0:
        param_dl = getattr(designer, 'param_dl', 0) or 0
        param_ul = max(designer.param_switch_ports - param_dl, 0)
        convergence['param'] = _conv_to_dict(calc_convergence_ratio(
            'param', param_dl, param_ul,
            _parse_speed_gbps(designer.param_speed),
            designer.param_leaf_count,
        ))
    if designer.storage_leaf_count > 0:
        storage_dl = getattr(designer, 'storage_dl', 0) or 0
        storage_ul = max(designer.storage_switch_ports - storage_dl, 0)
        convergence['storage'] = _conv_to_dict(calc_convergence_ratio(
            'storage', storage_dl, storage_ul,
            _parse_speed_gbps(designer.storage_speed),
            designer.storage_leaf_count,
        ))
    if getattr(designer, 'biz_enabled', True) and getattr(designer, 'biz_access', None):
        biz_ports = getattr(designer, 'biz_access_ports', 48)
        biz_uplinks = getattr(designer, 'biz_access_uplinks', 8)
        biz_speed = _parse_speed_gbps(getattr(designer, 'biz_port_speed', '25G'))
        convergence['biz'] = _conv_to_dict(calc_convergence_ratio(
            'biz', biz_ports, biz_uplinks, biz_speed, len(designer.biz_access),
        ))
    return convergence


def _get_iface_weight(port_str):
    """获取服务器接口类型排序权重: 参数网卡=1, 存储=2, OOB=3, 业务=4"""
    if '参数' in port_str:
        return 1
    if '存储' in port_str:
        return 2
    if 'OOB' in port_str:
        return 3
    if '业务' in port_str:
        return 4
    return 5


def _get_switch_type_weight(device_name):
    """获取交换机类型权重用于排序: Leaf=1, Spine=2, Core=3, 接入=4, 汇聚=5"""
    if "Leaf" in device_name:
        return 1
    elif "Spine" in device_name:
        return 2
    elif "Core" in device_name:
        return 3
    elif "接入" in device_name:
        return 4
    elif "汇聚" in device_name:
        return 5
    return 6


def generate_server_view(designer):
    """生成服务器视角的连接表"""
    connections = []

    for server in designer.servers:
        server_group = designer.server_groups[server.name]
        podid = designer.podid_map.get(server.name, "")
        for conn in server.connections:
            if conn.a_device == server.name:
                connections.append({
                    'podid': podid,
                    '服务器分组': server_group,
                    'A端设备': conn.a_device,
                    'A端接口': conn.a_port,
                    'A端模块': conn.a_module,
                    'A端机柜编号': conn.a_cabinet_name or '',
                    'A端U位': f"{conn.a_start_u}-{conn.a_end_u}" if conn.a_start_u else '',
                    'Z端设备': conn.z_device,
                    'Z端接口': conn.z_port,
                    'Z端模块': conn.z_module,
                    'Z端机柜编号': conn.z_cabinet_name or '',
                    'Z端U位': f"{conn.z_start_u}-{conn.z_end_u}" if conn.z_start_u else '',
                    '线缆类型': conn.cable_type,
                    '描述': conn.description
                })

    df = pd.DataFrame(connections)
    if not df.empty:
        # 按设备聚合, 同设备内按接口类型排序: 参数→存储→OOB→业务
        df['iface_w'] = df['A端接口'].apply(_get_iface_weight)
        df['port_num'] = df['A端接口'].apply(_extract_number)
        # 从设备名末尾提取数字用于自然排序 (GPU服务器_2 < GPU服务器_10)
        df['dev_num'] = df['A端设备'].apply(_extract_number)
        df = df.sort_values(by=['服务器分组', 'dev_num', 'iface_w', 'port_num'])
        df = df.drop(columns=['dev_num', 'iface_w', 'port_num'])
    return df


def generate_switch_view(designer):
    """生成交换机视角的连接表"""
    param_connections = []
    storage_connections = []

    all_switches = (designer.param_leaves + designer.param_spines + designer.param_cores +
                   designer.storage_leaves + designer.storage_spines + designer.storage_cores)

    for switch in all_switches:
        switch_group = designer.switch_groups.get(switch.name, "")
        podid = designer.podid_map.get(switch.name, "")
        for conn in switch.connections:
            if conn.a_device == switch.name:
                row = {
                    'podid': podid,
                    '交换机分组': switch_group,
                    'A端设备': conn.a_device,
                    'A端接口': conn.a_port,
                    'A端模块': conn.a_module,
                    'A端机柜编号': conn.a_cabinet_name or '',
                    'A端U位': f"{conn.a_start_u}-{conn.a_end_u}" if conn.a_start_u else '',
                    'Z端设备': conn.z_device,
                    'Z端接口': conn.z_port,
                    'Z端模块': conn.z_module,
                    'Z端机柜编号': conn.z_cabinet_name or '',
                    'Z端U位': f"{conn.z_start_u}-{conn.z_end_u}" if conn.z_start_u else '',
                    '线缆类型': conn.cable_type,
                    '描述': conn.description
                }
                if "参数" in switch.name:
                    param_connections.append(row)
                else:
                    storage_connections.append(row)

    param_df = pd.DataFrame(param_connections)
    storage_df = pd.DataFrame(storage_connections)

    if not param_df.empty:
        param_df['type_weight'] = param_df['A端设备'].apply(_get_switch_type_weight)
        param_df['dev_num'] = param_df['A端设备'].apply(_extract_number)
        param_df['port_num'] = param_df['A端接口'].apply(_extract_number)
        param_df = param_df.sort_values(by=['type_weight', '交换机分组', 'dev_num', 'A端设备', 'port_num'])
        param_df = param_df.drop(columns=['type_weight', 'dev_num', 'port_num'])

    if not storage_df.empty:
        storage_df['type_weight'] = storage_df['A端设备'].apply(_get_switch_type_weight)
        storage_df['dev_num'] = storage_df['A端设备'].apply(_extract_number)
        storage_df['port_num'] = storage_df['A端接口'].apply(_extract_number)
        storage_df = storage_df.sort_values(by=['type_weight', '交换机分组', 'dev_num', 'A端设备', 'port_num'])
        storage_df = storage_df.drop(columns=['type_weight', 'dev_num', 'port_num'])

    return {'参数网络': param_df, '存储网络': storage_df}


# V2.9.1: 机柜类型显示标签 (与 rack_allocation.CABINET_TYPE_* 对应)
_RACK_TYPE_LABELS = {'gpu': 'GPU柜', 'compute': '通算柜', 'storage': '存储柜', 'network': '网络柜',
                     'scaleup': 'Scale-Up柜'}  # V2.9.3-T4


def generate_summary_data(designer):
    """生成网络设计摘要数据"""
    half_ports = designer.param_switch_ports // 2
    if designer.param_3tier_needed:
        servers_per_group = min(designer.param_servers_per_pod // designer.param_ports_per_server, half_ports)
        param_downlink_usage = f"{servers_per_group}/{half_ports} ({servers_per_group / half_ports * 100:.1f}%)"
        param_uplink_usage = f"{half_ports}/{half_ports} (100.0%)"
    else:
        param_downlink_usage = f"{designer.param_servers_per_group}/{half_ports} ({designer.param_servers_per_group / half_ports * 100:.1f}%)"
        uplinks_used = min(half_ports, designer.param_spine_count)
        param_uplink_usage = f"{uplinks_used}/{half_ports} ({uplinks_used / half_ports * 100:.1f}%)"

    half_ports = designer.storage_switch_ports // 2
    if designer.storage_3tier_needed:
        servers_per_group = min(designer.storage_servers_per_pod // designer.storage_ports_per_server, half_ports)
        storage_downlink_usage = f"{servers_per_group}/{half_ports} ({servers_per_group / half_ports * 100:.1f}%)"
        storage_uplink_usage = f"{half_ports}/{half_ports} (100.0%)"
    else:
        storage_downlink_usage = f"{designer.storage_servers_per_group}/{half_ports} ({designer.storage_servers_per_group / half_ports * 100:.1f}%)"
        uplinks_used = min(half_ports, designer.storage_spine_count)
        storage_uplink_usage = f"{uplinks_used}/{half_ports} ({uplinks_used / half_ports * 100:.1f}%)"

    param_tier_info = "3层(Leaf-Spine-Core)" if designer.param_3tier_needed else "2层(Leaf-Spine)"
    storage_tier_info = "3层(Leaf-Spine-Core)" if designer.storage_3tier_needed else "2层(Leaf-Spine)"

    # V2.9.3-T6: 收敛比读 estimation 计算值, 不再硬编码
    convergence = _compute_convergence(designer)
    param_conv = convergence.get('param', {})
    storage_conv = convergence.get('storage', {})
    param_conv_str = f"1:{param_conv['convergenceRatio']:.2f}" if param_conv else ("1:1:1" if designer.param_3tier_needed else "1:1")
    storage_conv_str = f"1:{storage_conv['convergenceRatio']:.2f}" if storage_conv else ("1:1:1" if designer.storage_3tier_needed else "1:1")

    if designer.param_3tier_needed:
        param_group_info = f"{designer.param_pods}个POD, 每个POD{designer.param_servers_per_pod}台"
    else:
        param_group_info = f"{designer.param_groups}组, 每组{designer.param_servers_per_group}台"

    if designer.storage_3tier_needed:
        storage_group_info = f"{designer.storage_pods}个POD, 每个POD{designer.storage_servers_per_pod}台"
    else:
        storage_group_info = f"{designer.storage_groups}组, 每组{designer.storage_servers_per_group}台"

    summary = [
        ["网络设计摘要", ""],
        ["设计时间", pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["", ""],
        ["GPU服务器配置", ""],
        ["服务器数量", designer.num_servers],
        ["参数网卡/服务器", designer.param_ports_per_server],
        ["存储网卡/服务器", designer.storage_ports_per_server],
        ["", ""],
        ["参数网络设计", ""],
        ["交换机端口数", designer.param_switch_ports],
        ["Leaf交换机数量", designer.param_leaf_count],
        ["Spine交换机数量", designer.param_spine_count],
        ["Core交换机数量", designer.param_core_count if designer.param_3tier_needed else "无"],
        ["网络层级", param_tier_info],
        ["服务器分组", param_group_info],
        ["下行端口使用率", param_downlink_usage],
        ["上行端口使用率", param_uplink_usage],
        ["收敛比例", param_conv_str],
        ["", ""],
        ["存储网络设计", ""],
        ["交换机端口数", designer.storage_switch_ports],
        ["Leaf交换机数量", designer.storage_leaf_count],
        ["Spine交换机数量", designer.storage_spine_count],
        ["Core交换机数量", designer.storage_core_count if designer.storage_3tier_needed else "无"],
        ["网络层级", storage_tier_info],
        ["服务器分组", storage_group_info],
        ["下行端口使用率", storage_downlink_usage],
        ["上行端口使用率", storage_uplink_usage],
        ["收敛比例", storage_conv_str],
        ["", ""],
        ["网络速度配置", ""],
        ["参数网络速度", designer.param_speed],
        ["存储网络速度", designer.storage_speed],
        ["线缆类型", "参数: MPO | 存储: AOC | OOB: 网线/光纤 | 业务: 光纤"]
    ]
    return pd.DataFrame(summary, columns=["项目", "值"])


def generate_device_list(designer):
    """Generate device inventory from topology and device profiles

    V2.7.4-T9: 从 device_library 拉取 vendor/model 填充设备清单
    """
    from device_library import get_device_library

    items = []

    try:
        library = get_device_library()
    except Exception:
        library = None

    def _profile_id(profile):
        """从 device_profile 提取可哈希的 ID 字符串
        V2.7.4-T9fix: device_profile 可能是 LibraryDevice 对象或字符串 ID
        """
        if profile is None:
            return None
        # LibraryDevice 对象优先取 id 属性
        pid = getattr(profile, 'id', None)
        if pid:
            return pid
        # 字符串 ID 直接返回
        if isinstance(profile, str):
            return profile
        return None

    def _lookup(profile):
        """从设备库查询 vendor/model，返回 (vendor, model)"""
        if library and profile:
            pid = _profile_id(profile)
            if pid:
                dev = library.get(pid)
                if dev:
                    return getattr(dev, 'vendor', '') or '', getattr(dev, 'model', '') or ''
            # 直接传入对象时尝试从对象本身取 vendor/model
            v = getattr(profile, 'vendor', None)
            m = getattr(profile, 'model', None)
            if v or m:
                return v or '', m or ''
        return '', ''

    # Collect all device types and their counts from the designer
    # Servers: from designer.servers, group by group (e.g., "GPU服务器", "存储服务器")
    # V2.7.4-T9: 按 device_profile 分组以填充厂商型号
    server_groups = {}
    for server in designer.servers:
        group = server.group or "GPU服务器"
        profile = getattr(server, 'device_profile', None)
        pid = _profile_id(profile)
        group_key = (group, pid)
        if group_key not in server_groups:
            vendor, model = _lookup(profile)
            server_groups[group_key] = {
                "count": 0, "u_height": server.u_height or 4, "power": server.power_watts or 2000,
                "vendor": vendor, "model": model, "group": group,
            }
        server_groups[group_key]["count"] += 1

    for info in server_groups.values():
        items.append({
            "设备类型": info["group"],
            "厂商": info["vendor"],
            "型号": info["model"],
            "数量": info["count"],
            "单机功耗(W)": info["power"],
            "U位高度": info["u_height"],
            "总功耗(W)": info["count"] * info["power"],
            "总U位": info["count"] * info["u_height"],
        })

    # Switches: count by type prefix + device_profile
    # V2.7.4-T9: 按 (label, device_profile) 分组以填充厂商型号
    switch_types = {}
    all_switches = (designer.param_leaves + designer.param_spines + designer.param_cores +
                    designer.storage_leaves + designer.storage_spines + designer.storage_cores +
                    designer.oob_access + designer.oob_agg + designer.biz_access + designer.biz_agg)

    for sw in all_switches:
        stype = sw.obj_type.replace('param_', '').replace('storage_', '')
        stype_map = {
            'leaf': 'Leaf交换机', 'spine': 'Spine交换机', 'core': 'Core交换机',
            'access': '接入交换机', 'agg': '汇聚交换机',
        }
        label = stype_map.get(stype, stype)
        profile = getattr(sw, 'device_profile', None)
        pid = _profile_id(profile)
        sw_key = (label, pid)
        if sw_key not in switch_types:
            vendor, model = _lookup(profile)
            switch_types[sw_key] = {
                "count": 0, "power": sw.power_watts or 0, "u_height": sw.u_height or 1,
                "vendor": vendor, "model": model, "label": label,
            }
        switch_types[sw_key]["count"] += 1

    for info in switch_types.values():
        items.append({
            "设备类型": info["label"],
            "厂商": info["vendor"],
            "型号": info["model"],
            "数量": info["count"],
            "单机功耗(W)": info["power"],
            "U位高度": info["u_height"],
            "总功耗(W)": info["count"] * info["power"],
            "总U位": info["count"] * info["u_height"],
        })

    # Add totals row
    if len(items) > 0:
        total_power = sum(i["总功耗(W)"] for i in items)
        total_u = sum(i["总U位"] for i in items)
        items.append({
            "设备类型": "合计",
            "厂商": "",
            "型号": "",
            "数量": sum(i["数量"] for i in items),
            "单机功耗(W)": "",
            "U位高度": "",
            "总功耗(W)": total_power,
            "总U位": total_u,
        })

    columns = ["设备类型", "厂商", "型号", "数量", "单机功耗(W)", "U位高度", "总功耗(W)", "总U位"]
    return pd.DataFrame(items, columns=columns)


def apply_excel_formatting(filename):
    """应用Excel格式美化：合并同设备行、组分隔线、边框、交替行色"""
    from openpyxl.styles import Border, Side, PatternFill, Font, Alignment

    wb = load_workbook(filename)
    thin = Side(style='thin')
    thick = Side(style='medium')
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    thick_bottom = Border(left=thin, right=thin, top=thin, bottom=thick)
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(bold=True, size=12, color="FFFFFF")
    group_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    even_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    odd_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    col_widths = {
        'podid': 15, '服务器分组': 15, '交换机分组': 15,
        'A端设备': 30, 'A端接口': 15, 'A端模块': 12,
        'A端机柜编号': 15, 'A端U位': 12,
        'Z端设备': 30, 'Z端接口': 15, 'Z端模块': 12,
        'Z端机柜编号': 15, 'Z端U位': 12,
        '线缆类型': 12, '描述': 35, '项目': 25, '值': 40
    }

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        max_row = ws.max_row
        max_col = ws.max_column

        if max_row <= 1:
            continue

        ws.freeze_panes = 'A3'
        ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"

        # 列宽
        for col_idx, col_name in enumerate(ws[1], 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_name.value, 15)

        ws.row_dimensions[1].height = 28

        # 标题行
        for c in ws[1]:
            c.font = header_font; c.fill = header_fill
            c.alignment = center_align; c.border = thin_border

        # 识别列：从标题行确定podid_col, group_col, device_col, port_col
        headers = {cell.value: cell.column for cell in ws[1] if cell.value}
        podid_col = headers.get('podid')
        group_col = headers.get('服务器分组') or headers.get('交换机分组')
        device_col = headers.get('A端设备')
        port_col = headers.get('A端接口')

        # 摘要sheet特殊处理
        if sheet_name in ('网络设计摘要', '设计摘要'):
            for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
                for cell in row:
                    cell.alignment = center_align
                    cell.border = thin_border
            continue

        # V2.7.4-T11: 单次遍历完成 ffill + 合并区域收集 + 组分界标记
        # 旧实现需要 8 次遍历 (read→ffill→writeback→podid_merge→group_merge→device_merge→style→port_align)
        # 新实现合并为 2 次遍历 (collect+style)，万级连接导出 ≤ 5s
        merge_regions = []
        group_boundaries = set()
        ffill_cols = [c for c in [podid_col, group_col, device_col] if c]

        if max_row >= 2:
            # 单次遍历: 读取值 + ffill + 收集合并区域
            prev_vals = {c: None for c in ffill_cols}
            # 合并区域起点跟踪
            merge_start = {c: 2 for c in ffill_cols}
            dev_in_group_start = 2  # 设备列在当前分组内的起点

            for r in range(2, max_row + 1):
                # 读取当前行关键列的值
                cur_vals = {}
                for col in ffill_cols:
                    v = ws.cell(row=r, column=col).value
                    if v is None and prev_vals.get(col) is not None:
                        # NaN → ffill: 写回上一行的值
                        v = prev_vals[col]
                        ws.cell(row=r, column=col).value = v
                    cur_vals[col] = v

                # 检测 group 变更 (决定组分界 + 设备合并重置)
                grp_changed = (group_col and r > 2 and
                               cur_vals.get(group_col) != prev_vals.get(group_col))

                # group 变更时: 先关闭上一组的设备合并 + 标记组分界 + 关闭 group 合并
                if grp_changed:
                    # 关闭上一组的设备合并
                    if device_col and r - 1 > dev_in_group_start:
                        merge_regions.append((dev_in_group_start, r - 1, device_col, device_col))
                    dev_in_group_start = r
                    group_boundaries.add(r - 1)
                    # 关闭 group 合并
                    if group_col and r - 1 > merge_start[group_col]:
                        merge_regions.append((merge_start[group_col], r - 1, group_col, group_col))
                    merge_start[group_col] = r
                else:
                    # 同组内检测各列值变更以收集合并区域
                    for col in ffill_cols:
                        if col == group_col:
                            continue  # group_col 已在 grp_changed 分支处理
                        if col == device_col:
                            # 设备列在分组内独立合并
                            if cur_vals.get(device_col) != prev_vals.get(device_col) and r > 2:
                                if r - 1 > dev_in_group_start:
                                    merge_regions.append((dev_in_group_start, r - 1, device_col, device_col))
                                dev_in_group_start = r
                        else:
                            # podid 等普通列: 值变更即关闭合并
                            if cur_vals.get(col) != prev_vals.get(col) and r > 2:
                                if r - 1 > merge_start[col]:
                                    merge_regions.append((merge_start[col], r - 1, col, col))
                                merge_start[col] = r

                prev_vals = cur_vals

            # 收尾: 关闭所有未闭合的合并区域
            for col in ffill_cols:
                if col == device_col:
                    if max_row > dev_in_group_start:
                        merge_regions.append((dev_in_group_start, max_row, device_col, device_col))
                else:
                    if max_row > merge_start[col]:
                        merge_regions.append((merge_start[col], max_row, col, col))

        # 应用合并
        for sr, er, sc, ec in merge_regions:
            ws.merge_cells(start_row=sr, end_row=er, start_column=sc, end_column=ec)

        # ===== 数据行样式 (单次遍历, 含端口列对齐) =====
        for row_idx in range(2, max_row + 1):
            is_boundary = row_idx in group_boundaries
            fill = even_fill if row_idx % 2 == 0 else odd_fill
            for c in range(1, max_col + 1):
                cell = ws.cell(row=row_idx, column=c)
                cell.border = thick_bottom if is_boundary else thin_border
                if is_boundary and c == group_col:
                    cell.fill = group_fill
                else:
                    cell.fill = fill
                # 对齐: podid/group/device/端口列居中, 最后一列居左, 其余居中
                if c == max_col:
                    cell.alignment = left_align
                else:
                    cell.alignment = center_align

    wb.save(filename)
    print(f"已应用Excel格式美化: {filename}")


def export_all_connections(designer, filename):
    """导出所有连接关系到单个Excel文件"""
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        try:
            generate_summary_data(designer).to_excel(writer, sheet_name='网络设计摘要', index=False)
            generate_server_view(designer).to_excel(writer, sheet_name='服务器连接表', index=False)
            switch_views = generate_switch_view(designer)
            switch_views['参数网络'].to_excel(writer, sheet_name='参数网络连接表', index=False)
            switch_views['存储网络'].to_excel(writer, sheet_name='存储网络连接表', index=False)

            # OOB网络
            if designer.oob_enabled and designer.oob_access:
                oob_df = _generate_access_agg_view(designer, designer.oob_access, designer.oob_agg)
                oob_df.to_excel(writer, sheet_name='OOB网络连接表', index=False)
                # OOB汇聚反向视角
                oob_rev = _generate_agg_reverse_view(designer, designer.oob_access, designer.oob_agg, "OOB")
                if oob_rev is not None:
                    oob_rev.to_excel(writer, sheet_name='OOB汇聚视角', index=False)

            # 业务网络
            if designer.biz_enabled and designer.biz_access:
                biz_df = _generate_access_agg_view(designer, designer.biz_access, designer.biz_agg)
                biz_df.to_excel(writer, sheet_name='业务网络连接表', index=False)

        except Exception as e:
            print(f"生成连接表时出错: {e}")
            if len(writer.book.sheetnames) == 0:
                writer.book.create_sheet("错误信息")
            raise

    apply_excel_formatting(filename)
    print(f"所有连接表已导出到: {filename}")


def _generate_access_agg_view(designer, access_switches, agg_switches):
    """生成接入-汇聚网络交换机视角的连接表"""
    connections = []
    all_switches = access_switches + agg_switches

    for sw in all_switches:
        sw_group = designer.switch_groups.get(sw.name, "")
        podid = designer.podid_map.get(sw.name, "")
        for conn in sw.connections:
            if conn.a_device == sw.name:
                connections.append({
                    'podid': podid,
                    '交换机分组': sw_group,
                    'A端设备': conn.a_device,
                    'A端接口': conn.a_port,
                    'A端模块': conn.a_module,
                    'A端机柜编号': conn.a_cabinet_name or '',
                    'A端U位': f"{conn.a_start_u}-{conn.a_end_u}" if conn.a_start_u else '',
                    'Z端设备': conn.z_device,
                    'Z端接口': conn.z_port,
                    'Z端模块': conn.z_module,
                    'Z端机柜编号': conn.z_cabinet_name or '',
                    'Z端U位': f"{conn.z_start_u}-{conn.z_end_u}" if conn.z_start_u else '',
                    '线缆类型': conn.cable_type,
                    '描述': conn.description
                })

    df = pd.DataFrame(connections)
    if not df.empty:
        df['type_weight'] = df['A端设备'].apply(_get_switch_type_weight)
        df['dev_num'] = df['A端设备'].apply(_extract_number)
        df['port_num'] = df['A端接口'].apply(_extract_number)
        df = df.sort_values(by=['type_weight', '交换机分组', 'dev_num', 'A端设备', 'port_num'])
        df = df.drop(columns=['type_weight', 'dev_num', 'port_num'])
    return df


def _generate_agg_reverse_view(designer, access_switches, agg_switches, network_name):
    """生成汇聚交换机反向视角 (A端=汇聚, Z端=接入)"""
    connections = []

    for agg in agg_switches:
        sw_group = designer.switch_groups.get(agg.name, "")
        podid = designer.podid_map.get(agg.name, "")
        for conn in agg.connections:
            if conn.a_device == agg.name:
                connections.append({
                    'podid': podid,
                    '交换机分组': sw_group,
                    'A端设备': conn.a_device,
                    'A端接口': conn.a_port,
                    'A端模块': conn.a_module,
                    'A端机柜编号': conn.a_cabinet_name or '',
                    'A端U位': f"{conn.a_start_u}-{conn.a_end_u}" if conn.a_start_u else '',
                    'Z端设备': conn.z_device,
                    'Z端接口': conn.z_port,
                    'Z端模块': conn.z_module,
                    'Z端机柜编号': conn.z_cabinet_name or '',
                    'Z端U位': f"{conn.z_start_u}-{conn.z_end_u}" if conn.z_start_u else '',
                    '线缆类型': conn.cable_type,
                    '描述': conn.description
                })

    df = pd.DataFrame(connections)
    if not df.empty:
        df['dev_num'] = df['A端设备'].apply(_extract_number)
        df['port_num'] = df['A端接口'].apply(_extract_number)
        df = df.sort_values(by=['交换机分组', 'dev_num', 'A端设备', 'port_num'])
        df = df.drop(columns=['dev_num', 'port_num'])
    return df if not df.empty else None


# ================================================================
# V2.4 新增：布线指导表 / BOM / 报告数据
# ================================================================

def export_cabling_guide(designer, filename):
    """V2.4: 导出布线指导表（含光模块型号、长度估算、成本）

    每条连接生成一行，包含：
    - A/Z 端设备、端口、机柜、U位
    - 线缆类型、速率
    - 光模块型号、规格、封装、光纤类型
    - 估算长度、价格区间
    """
    from optical_selector import select_module_for_connection, estimate_module_cost

    rows = []
    all_switches = (
        designer.param_leaves + designer.param_spines + designer.param_cores +
        designer.storage_leaves + designer.storage_spines + designer.storage_cores +
        designer.oob_access + designer.oob_agg + designer.biz_access + designer.biz_agg
    )
    all_devices = designer.servers + all_switches
    seen_conns = set()

    for dev in all_devices:
        for conn in dev.connections:
            if conn.a_device != dev.name:
                continue
            # 去重（每条连接在 A 端和 Z 端各存一份）
            pair_key = tuple(sorted([conn.a_device, conn.z_device])) + (conn.a_port,)
            if pair_key in seen_conns:
                continue
            seen_conns.add(pair_key)

            sel = select_module_for_connection(conn)
            cost_lo, cost_hi = estimate_module_cost(sel.price_range) if sel else (0, 0)

            rows.append({
                '网络类型': conn.network_type or '',
                'A端设备': conn.a_device,
                'A端端口': conn.a_port,
                'A端机柜': conn.a_cabinet_name or '',
                'A端U位': f"{conn.a_start_u}-{conn.a_end_u}" if conn.a_start_u else '',
                'Z端设备': conn.z_device,
                'Z端端口': conn.z_port,
                'Z端机柜': conn.z_cabinet_name or '',
                'Z端U位': f"{conn.z_start_u}-{conn.z_end_u}" if conn.z_start_u else '',
                '速率': conn.a_module or '',
                '线缆类型': conn.cable_type,
                '光模块型号': sel.module_id if sel else '',
                '封装': sel.form_factor if sel else '',
                '规格': sel.spec if sel else '',
                '光纤类型': sel.fiber_type if sel else '',
                '支持距离(m)': sel.distance_m if sel else '',
                '估算长度(m)': f"{sel.estimated_length_m:.1f}" if sel else '',
                '价格区间': sel.price_range if sel else '',
                '估价低(元)': cost_lo,
                '估价高(元)': cost_hi,
                '描述': conn.description,
            })

    df = pd.DataFrame(rows)
    if not df.empty:
        df['__num'] = df['A端设备'].apply(_extract_number)
        df = df.sort_values(by=['网络类型', '__num', 'A端设备'])
        df = df.drop(columns=['__num'])

    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='布线指导表', index=False)
        # 汇总 sheet
        if not df.empty:
            summary = df.groupby(['速率', '光模块型号', '价格区间']).agg(
                数量=('A端设备', 'count'),
                估价低合计=('估价低(元)', 'sum'),
                估价高合计=('估价高(元)', 'sum'),
            ).reset_index()
            summary.to_excel(writer, sheet_name='光模块汇总', index=False)

    apply_excel_formatting(filename)
    print(f"布线指导表已导出: {filename}")
    return df


def export_bom(designer, filename):
    """V2.4: 导出 BOM 成本估算表

    汇总所有设备 + 光模块，按型号分组，含数量和价格区间
    """
    from optical_selector import select_module_for_connection, estimate_module_cost, PRICE_RANGE_MAP, LEAD_TIME_MAP
    from device_library import get_device_library

    try:
        library = get_device_library()
    except Exception:
        library = None

    rows = []

    # 1. 服务器 (V2.9.3-T6: 按型号聚合, 替代逐台一行)
    server_agg = {}
    for server in designer.servers:
        dev = None
        if library and getattr(server, 'device_profile', None):
            # V2.9.3: 传入 id 而非 LibraryDevice 对象(对象不可哈希导致 TypeError)
            dev = library.get(getattr(server.device_profile, 'id', '')) or None
        price = getattr(dev, 'price_range', None) if dev else None
        category = 'GPU服务器' if 'GPU' in (server.name + getattr(dev, 'description', '') or '') else '服务器'
        model = getattr(dev, 'id', '') if dev else ''
        desc = getattr(dev, 'description', '') if dev else server.obj_type
        lead = LEAD_TIME_MAP.get(price or '', '')
        key = (category, model, desc, price, lead)
        if key not in server_agg:
            server_agg[key] = {'power': server.power_watts or 0, 'count': 0}
        server_agg[key]['count'] += 1
    for (category, model, desc, price, lead), info in server_agg.items():
        rows.append({
            '类别': category,
            '设备名称': model or desc,
            '设备型号': model,
            '描述': desc,
            '数量': info['count'],
            '单位功率(W)': info['power'],
            '价格区间': price or '',
            '供货周期': lead,
        })

    # 2. 交换机 (V2.9.3-T6: 按型号聚合)
    all_switches = (
        designer.param_leaves + designer.param_spines + designer.param_cores +
        designer.storage_leaves + designer.storage_spines + designer.storage_cores +
        designer.oob_access + designer.oob_agg + designer.biz_access + designer.biz_agg
    )
    switch_agg = {}
    for sw in all_switches:
        dev = None
        if library and getattr(sw, 'device_profile', None):
            # V2.9.3: 传入 id 而非 LibraryDevice 对象(对象不可哈希导致 TypeError)
            dev = library.get(getattr(sw.device_profile, 'id', '')) or None
        price = getattr(dev, 'price_range', None) if dev else None
        sw_type = '参数网交换机' if '参数' in sw.name else \
                  '存储网交换机' if '存储' in sw.name else \
                  'OOB交换机' if 'OOB' in sw.name else \
                  '业务交换机' if '业务' in sw.name else '交换机'
        model = getattr(dev, 'id', '') if dev else ''
        desc = getattr(dev, 'description', '') if dev else sw.obj_type
        lead = LEAD_TIME_MAP.get(price or '', '')
        key = (sw_type, model, desc, price, lead)
        if key not in switch_agg:
            switch_agg[key] = {'power': sw.power_watts or 0, 'count': 0}
        switch_agg[key]['count'] += 1
    for (sw_type, model, desc, price, lead), info in switch_agg.items():
        rows.append({
            '类别': sw_type,
            '设备名称': model or desc,
            '设备型号': model,
            '描述': desc,
            '数量': info['count'],
            '单位功率(W)': info['power'],
            '价格区间': price or '',
            '供货周期': lead,
        })

    # 3. 光模块（按型号汇总）
    # V2.7.4-T3: 增加功耗和供货周期
    seen_conns = set()
    module_counts = {}
    for dev in designer.servers + all_switches:
        for conn in dev.connections:
            if conn.a_device != dev.name:
                continue
            pair_key = tuple(sorted([conn.a_device, conn.z_device])) + (conn.a_port,)
            if pair_key in seen_conns:
                continue
            seen_conns.add(pair_key)

            sel = select_module_for_connection(conn, library)
            if sel:
                key = sel.module_id
                if key not in module_counts:
                    module_counts[key] = {
                        'category': '光模块',
                        'name': sel.module_id,
                        'model': sel.module_id,
                        'desc': sel.description,
                        'count': 0,
                        'power': sel.power_w,  # V2.7.4-T3: 真实功耗
                        'price': sel.price_range,
                        'lead_time': sel.lead_time_weeks,  # V2.7.4-T3: 供货周期
                    }
                module_counts[key]['count'] += 1

    for mod_data in module_counts.values():
        rows.append({
            '类别': '光模块',
            '设备名称': mod_data['name'],
            '设备型号': mod_data['model'],
            '描述': mod_data['desc'],
            '数量': mod_data['count'],
            '单位功率(W)': mod_data['power'],
            '价格区间': mod_data['price'],
            '供货周期': mod_data['lead_time'],
        })

    df = pd.DataFrame(rows)

    # V2.7.4-T3: 计算价格估算（统一引用 optical_selector.PRICE_RANGE_MAP，消除重复）
    df['估价低(元)'] = df['价格区间'].map(lambda x: PRICE_RANGE_MAP.get(x, (0, 0))[0])
    df['估价高(元)'] = df['价格区间'].map(lambda x: PRICE_RANGE_MAP.get(x, (0, 0))[1])
    df['估价低小计'] = df['估价低(元)'] * df['数量']
    df['估价高小计'] = df['估价高(元)'] * df['数量']

    # 汇总
    total_lo = df['估价低小计'].sum()
    total_hi = df['估价高小计'].sum()
    total_power = (df['单位功率(W)'] * df['数量']).sum()

    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='BOM清单', index=False)
        # 按类别汇总
        cat_summary = df.groupby('类别').agg(
            数量=('数量', 'sum'),
            估价低合计=('估价低小计', 'sum'),
            估价高合计=('估价高小计', 'sum'),
            总功率W=('单位功率(W)', lambda x: (x * df.loc[x.index, '数量']).sum()),
        ).reset_index()
        cat_summary.to_excel(writer, sheet_name='类别汇总', index=False)
        # 总计
        total_df = pd.DataFrame([{
            '项目': '总计',
            '设备总数': df['数量'].sum(),
            '估价低(元)': total_lo,
            '估价高(元)': total_hi,
            '总功率(W)': total_power,
            '估价区间': f"¥{total_lo:,} ~ ¥{total_hi:,}",
        }])
        total_df.to_excel(writer, sheet_name='总计', index=False)

    apply_excel_formatting(filename)
    print(f"BOM已导出: {filename} (估价: ¥{total_lo:,} ~ ¥{total_hi:,})")
    return df


def generate_report_data(designer, estimation=None):
    """V2.4: 生成 PDF 报告所需的完整数据（字典格式）

    返回包含以下章节的数据：
    1. 项目概览
    2. 网络架构
    3. 设备清单
    4. 拓扑参数
    5. 功耗与散热
    6. 布线汇总
    7. 成本估算
    8. 校验结果

    V2.9.3-T6: 项目名称取自配置 meta.name; 收敛比优先读 estimation 值
    """
    from optical_selector import select_module_for_connection, estimate_module_cost

    # 1. 项目概览
    pc = getattr(designer, '_project_config', None) or {}
    project_name = pc.get('meta', {}).get('name', '') or 'AutoLink 项目'
    overview = {
        '项目名称': project_name,
        'GPU服务器数': designer.num_servers,
        '存储服务器数': designer.additional_storage,
        '通算服务器数': designer.additional_compute,
        '服务器总数': designer.total_servers,
        '参数网速率': designer.param_speed,
        '存储网速率': designer.storage_speed,
        '下行模式': designer.downlink_mode,
    }
    # V2.9.3-T4: Scale-Up 概览
    su_cfg = getattr(designer, 'scale_up_config', None)
    if su_cfg:
        su_stats = getattr(designer, 'scale_up_stats', {})
        overview['Scale-Up协议'] = su_cfg.get('protocol', '')
        overview['Scale-Up GPU节点数'] = len(getattr(designer, 'scale_up_gpus', []))
        overview['Scale-Up域数'] = su_stats.get('num_domains', 0)
        overview['Scale-Up总链路数'] = su_stats.get('total_connections', 0)

    # 2. 网络架构
    all_switches = (
        designer.param_leaves + designer.param_spines + designer.param_cores +
        designer.storage_leaves + designer.storage_spines + designer.storage_cores +
        designer.oob_access + designer.oob_agg + designer.biz_access + designer.biz_agg
    )
    architecture = {
        '参数网Leaf': len(designer.param_leaves),
        '参数网Spine': len(designer.param_spines),
        '参数网Core': len(designer.param_cores),
        '存储网Leaf': len(designer.storage_leaves),
        '存储网Spine': len(designer.storage_spines),
        '存储网Core': len(designer.storage_cores),
        'OOB接入': len(designer.oob_access),
        'OOB汇聚': len(designer.oob_agg),
        '业务接入': len(designer.biz_access),
        '业务汇聚': len(designer.biz_agg),
        '交换机总数': len(all_switches),
        # V2.9.3-T4: Scale-Up
        'Scale-Up GPU节点': len(getattr(designer, 'scale_up_gpus', [])),
    }

    # 3. 功耗
    server_power = sum(s.power_watts or 0 for s in designer.servers)
    switch_power = sum(sw.power_watts or 0 for sw in all_switches)
    total_power = server_power + switch_power
    power = {
        '服务器功耗(W)': server_power,
        '交换机功耗(W)': switch_power,
        '总IT功耗(W)': total_power,
        '总IT功耗(kW)': round(total_power / 1000, 2),
        '机柜功率限制(W)': designer.power_limit_per_rack,
        '散热方式': '风冷',
    }

    # 4. 校验结果
    validation = designer.validate_topology()

    # 5. 光模块汇总
    seen_conns = set()
    module_stats = {}
    for dev in designer.servers + all_switches:
        for conn in dev.connections:
            if conn.a_device != dev.name:
                continue
            pair_key = tuple(sorted([conn.a_device, conn.z_device])) + (conn.a_port,)
            if pair_key in seen_conns:
                continue
            seen_conns.add(pair_key)
            sel = select_module_for_connection(conn)
            if sel:
                if sel.module_id not in module_stats:
                    module_stats[sel.module_id] = {'count': 0, 'price': sel.price_range, 'spec': sel.spec}
                module_stats[sel.module_id]['count'] += 1

    # 6. 成本估算
    price_map = {'低': (500, 2000), '中': (2000, 8000), '高': (8000, 30000), '极高': (30000, 100000)}
    module_cost_lo = sum(price_map.get(s['price'], (0, 0))[0] * s['count'] for s in module_stats.values())
    module_cost_hi = sum(price_map.get(s['price'], (0, 0))[1] * s['count'] for s in module_stats.values())

    cost = {
        '光模块总数': sum(s['count'] for s in module_stats.values()),
        '光模块估价低(元)': module_cost_lo,
        '光模块估价高(元)': module_cost_hi,
        '光模块估价区间': f"¥{module_cost_lo:,} ~ ¥{module_cost_hi:,}",
    }

    # 7. 机柜规划 (V2.9.1: 机柜清单,含交换机; 类型来自 rack_allocation 分配)
    rack_type_map = {cab.id: cab.type for cab in (getattr(designer, '_rack_cabinets', []) or [])}
    rack_cabs = {}
    for dev in designer.servers + all_switches + getattr(designer, 'scale_up_gpus', []):
        if dev.cabinet_id is None:
            continue
        cid = dev.cabinet_id
        if cid not in rack_cabs:
            rack_cabs[cid] = {
                '柜号': dev.cabinet_name or f"机柜{cid}",
                '类型': _RACK_TYPE_LABELS.get(rack_type_map.get(cid, 'gpu'), rack_type_map.get(cid, 'gpu')),
                '设备数': 0,
                '总功率(W)': 0,
                '功率上限(W)': designer.power_limit_per_rack,
                '利用率(%)': 0,
                '超限': False,
                '设备': [],
            }
        rack_cabs[cid]['设备数'] += 1
        rack_cabs[cid]['总功率(W)'] += dev.power_watts or 0
        rack_cabs[cid]['设备'].append(
            f"{dev.name}({dev.power_watts or 0}W/U{dev.start_u}-U{dev.end_u})"
        )
    racks = []
    for cid in sorted(rack_cabs.keys()):
        c = rack_cabs[cid]
        c['利用率(%)'] = round(c['总功率(W)'] / c['功率上限(W)'] * 100, 1) if c['功率上限(W)'] else 0
        c['超限'] = c['总功率(W)'] > c['功率上限(W)']
        racks.append(c)

    return {
        'overview': overview,
        'architecture': architecture,
        'power': power,
        'validation': validation,
        'modules': module_stats,
        'cost': cost,
        'racks': racks,
        # V2.9.3-T6: 设备清单(按型号聚合) + 收敛比(优先读 estimation)
        'devices': generate_device_list(designer),
        'convergence': (estimation or {}).get('convergence', {}) if estimation else _compute_convergence(designer),
        'generated_at': pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S'),
    }


def export_compliance_report(designer, filename):
    """V2.7.5-T8: 导出信创合规报告

    统计国产/进口/未标注设备清单与占比，导出 Excel 报告。
    """
    from device_library import get_device_library

    try:
        library = get_device_library()
    except Exception:
        library = None

    def _get_origin(device_profile):
        """从设备档案获取 origin 字段"""
        if device_profile is None:
            return 'unknown'
        # LibraryDevice 对象
        origin = getattr(device_profile, 'origin', None)
        if origin:
            return origin
        # 尝试从 library 查询
        pid = getattr(device_profile, 'id', None)
        if pid and library:
            dev = library.get(pid)
            if dev and dev.origin:
                return dev.origin
        return 'unknown'

    def _origin_label(origin):
        return {'domestic': '国产', 'imported': '进口', 'mixed': '混合', 'unknown': '未标注'}.get(origin, '未标注')

    rows = []

    # 服务器
    for server in designer.servers:
        profile = getattr(server, 'device_profile', None)
        origin = _get_origin(profile)
        rows.append({
            '设备名称': server.name,
            '设备类型': getattr(server, 'group', '') or 'GPU服务器',
            '厂商': getattr(profile, 'vendor', '') if profile else '',
            '型号': getattr(profile, 'model', '') if profile else '',
            '属性': _origin_label(origin),
            'origin': origin,
            '数量': 1,
        })

    # 交换机
    all_switches = (designer.param_leaves + designer.param_spines + designer.param_cores +
                    designer.storage_leaves + designer.storage_spines + designer.storage_cores +
                    designer.oob_access + designer.oob_agg + designer.biz_access + designer.biz_agg)
    for sw in all_switches:
        profile = getattr(sw, 'device_profile', None)
        origin = _get_origin(profile)
        rows.append({
            '设备名称': sw.name,
            '设备类型': getattr(sw, 'obj_type', ''),
            '厂商': getattr(profile, 'vendor', '') if profile else '',
            '型号': getattr(profile, 'model', '') if profile else '',
            '属性': _origin_label(origin),
            'origin': origin,
            '数量': 1,
        })

    df = pd.DataFrame(rows)

    # 汇总统计
    total = len(df)
    domestic = len(df[df['origin'] == 'domestic'])
    imported = len(df[df['origin'] == 'imported'])
    unknown = len(df[df['origin'] == 'unknown'])
    domestic_ratio = domestic / total * 100 if total > 0 else 0

    summary = pd.DataFrame([
        ['信创合规报告', ''],
        ['生成时间', pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')],
        ['', ''],
        ['统计汇总', ''],
        ['设备总数', total],
        ['国产设备数', domestic],
        ['进口设备数', imported],
        ['未标注设备数', unknown],
        ['国产化率', f'{domestic_ratio:.1f}%'],
        ['信创合规', '达标 (≥50%)' if domestic_ratio >= 50 else '未达标 (<50%)'],
        ['', ''],
        ['说明', '国产化率 = 国产设备数 / 设备总数 × 100%'],
        ['合规标准', '国产化率 ≥ 50% 为信创合规达标'],
    ], columns=['项目', '值'])

    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        summary.to_excel(writer, sheet_name='信创汇总', index=False)
        if not df.empty:
            df_sorted = df.sort_values(by=['origin', '设备类型', '设备名称'])
            df_sorted.drop(columns=['origin']).to_excel(writer, sheet_name='设备清单', index=False)

    apply_excel_formatting(filename)
    print(f"已导出信创合规报告: {filename}")
    return {
        'total': total,
        'domestic': domestic,
        'imported': imported,
        'unknown': unknown,
        'domestic_ratio': round(domestic_ratio, 1),
        'compliant': domestic_ratio >= 50,
    }


def export_pdf_report(designer, filename):
    """V2.4.6: 生成 PDF 设计报告

    V2.7.4-T7: 增加功率分布柱状图和光模块成本饼图
    V2.7.4-T8: 增加页眉页脚和目录
    """
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, Image,
    )
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import os as _os
    import tempfile

    # 注册中文字体（Windows: 微软雅黑；Linux: 文泉驿；fallback: Helvetica）
    font_name = 'Helvetica'
    for font_path in [
        ('C:/Windows/Fonts/msyh.ttc', 'MSYH'),
        ('C:/Windows/Fonts/simhei.ttf', 'SimHei'),
        ('/usr/share/fonts/truetype/wqy/wqy-microhei.ttc', 'WQY'),
        ('/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf', 'DejaVu'),
    ]:
        if _os.path.exists(font_path[0]):
            try:
                pdfmetrics.registerFont(TTFont(font_path[1], font_path[0]))
                font_name = font_path[1]
                break
            except Exception:
                continue

    data = generate_report_data(designer)
    project_name = data.get('overview', {}).get('项目名称', 'AutoLink')

    # V2.7.4-T8: 页眉页脚回调
    def _draw_header_footer(canvas, doc):
        canvas.saveState()
        # 页眉：项目名 + 版本
        canvas.setFont(font_name, 8)
        canvas.setFillColor(colors.HexColor('#6b7280'))
        canvas.drawString(20 * mm, A4[1] - 12 * mm, f"AutoLink 智算中心设计报告 - {project_name}")
        canvas.drawRightString(A4[0] - 20 * mm, A4[1] - 12 * mm, f"v{data.get('version', '2.7.4')}")
        canvas.setStrokeColor(colors.HexColor('#d1d5db'))
        canvas.line(20 * mm, A4[1] - 14 * mm, A4[0] - 20 * mm, A4[1] - 14 * mm)
        # 页脚：页码
        canvas.drawCentredString(A4[0] / 2, 10 * mm, f"第 {doc.page} 页")
        canvas.restoreState()

    doc = SimpleDocTemplate(
        filename, pagesize=A4,
        leftMargin=20 * mm, rightMargin=20 * mm,
        topMargin=25 * mm, bottomMargin=20 * mm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('ChTitle', parent=styles['Title'],
                                 fontName=font_name, fontSize=20, spaceAfter=12)
    h2_style = ParagraphStyle('ChH2', parent=styles['Heading2'],
                              fontName=font_name, fontSize=14, spaceBefore=10, spaceAfter=6,
                              textColor=colors.HexColor('#1e40af'))
    normal_style = ParagraphStyle('ChNormal', parent=styles['Normal'],
                                  fontName=font_name, fontSize=10, leading=16)
    cell_style = ParagraphStyle('ChCell', parent=styles['Normal'],
                                fontName=font_name, fontSize=9, leading=12)
    toc_style = ParagraphStyle('ChTOC', parent=styles['Normal'],
                               fontName=font_name, fontSize=11, leading=20)

    story = []

    # V2.7.4-T8: 目录
    story.append(Paragraph('目录', h2_style))
    toc_items = ['1. 项目概览', '2. 网络架构', '3. 设备清单', '4. 收敛比',
                 '5. 功耗与散热', '6. 光模块汇总', '7. 成本估算', '8. 机柜规划', '9. 校验结果']
    for item in toc_items:
        story.append(Paragraph(item, toc_style))
    story.append(PageBreak())

    # 封面
    story.append(Paragraph('AutoLink 智算中心设计报告', title_style))
    story.append(Spacer(1, 6 * mm))
    story.append(Paragraph(f"生成时间：{data['generated_at']}", normal_style))
    story.append(Spacer(1, 4 * mm))

    # 1. 项目概览
    story.append(Paragraph('1. 项目概览', h2_style))
    ov = data['overview']
    ov_rows = [[Paragraph('项目', cell_style), Paragraph('数值', cell_style)]]
    for k, v in ov.items():
        ov_rows.append([Paragraph(str(k), cell_style), Paragraph(str(v), cell_style)])
    ov_table = Table(ov_rows, colWidths=[60 * mm, 80 * mm])
    ov_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f3f4f6')]),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
    ]))
    story.append(ov_table)
    story.append(Spacer(1, 4 * mm))

    # 2. 网络架构
    story.append(Paragraph('2. 网络架构', h2_style))
    arch = data['architecture']
    arch_rows = [[Paragraph('设备类型', cell_style), Paragraph('数量', cell_style)]]
    for k, v in arch.items():
        arch_rows.append([Paragraph(str(k), cell_style), Paragraph(str(v), cell_style)])
    arch_table = Table(arch_rows, colWidths=[60 * mm, 40 * mm])
    arch_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10b981')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f3f4f6')]),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    story.append(arch_table)
    story.append(Spacer(1, 4 * mm))

    # V2.9.3-T6: 3. 设备清单 (按型号聚合)
    story.append(Paragraph('3. 设备清单', h2_style))
    devices = data.get('devices', [])
    if isinstance(devices, pd.DataFrame):
        devices = devices.to_dict('records') if not devices.empty else []
    if devices:
        dev_rows = [[Paragraph('设备类型', cell_style), Paragraph('厂商', cell_style),
                     Paragraph('型号', cell_style), Paragraph('数量', cell_style),
                     Paragraph('单机功耗(W)', cell_style), Paragraph('U位', cell_style)]]
        for d in devices:
            dev_rows.append([
                Paragraph(str(d.get('设备类型', '')), cell_style),
                Paragraph(str(d.get('厂商', '')), cell_style),
                Paragraph(str(d.get('型号', '')), cell_style),
                Paragraph(str(d.get('数量', '')), cell_style),
                Paragraph(str(d.get('单机功耗(W)', '')), cell_style),
                Paragraph(str(d.get('U位高度', '')), cell_style),
            ])
        dev_table = Table(dev_rows, colWidths=[28 * mm, 24 * mm, 40 * mm, 14 * mm, 22 * mm, 14 * mm],
                          repeatRows=1)
        dev_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10b981')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f3f4f6')]),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        story.append(dev_table)
        story.append(Spacer(1, 4 * mm))
    else:
        story.append(Paragraph('无设备数据', normal_style))

    # V2.9.3-T6: 4. 收敛比 (读 estimation 计算值)
    story.append(Paragraph('4. 收敛比', h2_style))
    convergence = data.get('convergence', {})
    if convergence:
        conv_rows = [[Paragraph('网络', cell_style), Paragraph('收敛比', cell_style),
                      Paragraph('下行带宽(Gbps)', cell_style), Paragraph('上行带宽(Gbps)', cell_style),
                      Paragraph('状态', cell_style)]]
        net_labels = {'param': '参数网', 'storage': '存储网', 'biz': '业务网'}
        for net, info in convergence.items():
            ratio = info.get('convergenceRatio', 1)
            conv_rows.append([
                Paragraph(net_labels.get(net, net), cell_style),
                Paragraph(f"1:{ratio:.2f}", cell_style),
                Paragraph(str(info.get('downlinkBwGbps', '')), cell_style),
                Paragraph(str(info.get('uplinkBwGbps', '')), cell_style),
                Paragraph('无阻塞' if info.get('meetsTarget') else '有收敛', cell_style),
            ])
        conv_table = Table(conv_rows, colWidths=[24 * mm, 22 * mm, 32 * mm, 32 * mm, 22 * mm])
        conv_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0ea5e9')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f3f4f6')]),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        story.append(conv_table)
        story.append(Spacer(1, 4 * mm))
    else:
        story.append(Paragraph('无收敛比数据', normal_style))

    # 5. 功耗与散热
    story.append(Paragraph('5. 功耗与散热', h2_style))
    pw = data['power']
    pw_rows = [[Paragraph('项目', cell_style), Paragraph('数值', cell_style)]]
    for k, v in pw.items():
        pw_rows.append([Paragraph(str(k), cell_style), Paragraph(str(v), cell_style)])
    pw_table = Table(pw_rows, colWidths=[60 * mm, 60 * mm])
    pw_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f59e0b')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f3f4f6')]),
    ]))
    story.append(pw_table)
    story.append(Spacer(1, 4 * mm))

    # V2.7.4-T7: 功率分布柱状图
    power_chart_path = _generate_power_chart(pw, font_name)
    if power_chart_path:
        story.append(Paragraph('功率分布图：', normal_style))
        story.append(Image(power_chart_path, width=140 * mm, height=80 * mm))
        story.append(Spacer(1, 4 * mm))

    story.append(PageBreak())

    # 6. 光模块汇总
    story.append(Paragraph('6. 光模块汇总', h2_style))
    mods = data['modules']
    if mods:
        mod_rows = [[Paragraph('型号', cell_style), Paragraph('数量', cell_style),
                     Paragraph('规格', cell_style), Paragraph('价位', cell_style)]]
        for mid, info in mods.items():
            mod_rows.append([
                Paragraph(str(mid), cell_style),
                Paragraph(str(info['count']), cell_style),
                Paragraph(str(info.get('spec', '')) if isinstance(info.get('spec'), str) else '', cell_style),
                Paragraph(str(info.get('price', '')), cell_style),
            ])
        mod_table = Table(mod_rows, colWidths=[40 * mm, 20 * mm, 60 * mm, 20 * mm])
        mod_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#8b5cf6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f3f4f6')]),
        ]))
        story.append(mod_table)
        story.append(Spacer(1, 4 * mm))

        # V2.7.4-T7: 光模块成本饼图
        cost_chart_path = _generate_module_cost_chart(mods, font_name)
        if cost_chart_path:
            story.append(Paragraph('光模块成本占比图：', normal_style))
            story.append(Image(cost_chart_path, width=120 * mm, height=80 * mm))
    else:
        story.append(Paragraph('无光模块数据', normal_style))
    story.append(Spacer(1, 4 * mm))

    # 7. 成本估算
    story.append(Paragraph('7. 成本估算', h2_style))
    cost = data['cost']
    cost_rows = [[Paragraph('项目', cell_style), Paragraph('数值', cell_style)]]
    for k, v in cost.items():
        cost_rows.append([Paragraph(str(k), cell_style), Paragraph(str(v), cell_style)])
    cost_table = Table(cost_rows, colWidths=[60 * mm, 80 * mm])
    cost_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#ef4444')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f3f4f6')]),
    ]))
    story.append(cost_table)
    story.append(Spacer(1, 4 * mm))

    # V2.9.1: 8. 机柜规划（机柜清单，含交换机；V2.9.3-T6: 全量渲染不分页截断）
    story.append(Paragraph('8. 机柜规划', h2_style))
    racks = data.get('racks', [])
    if racks:
        rack_rows = [[Paragraph('柜号', cell_style), Paragraph('类型', cell_style),
                      Paragraph('设备数', cell_style), Paragraph('总功率(W)', cell_style),
                      Paragraph('上限(W)', cell_style), Paragraph('利用率', cell_style),
                      Paragraph('超限', cell_style)]]
        for r in racks:
            rack_rows.append([
                Paragraph(str(r['柜号']), cell_style),
                Paragraph(str(r['类型']), cell_style),
                Paragraph(str(r['设备数']), cell_style),
                Paragraph(str(r['总功率(W)']), cell_style),
                Paragraph(str(r['功率上限(W)']), cell_style),
                Paragraph(f"{r['利用率(%)']}%", cell_style),
                Paragraph('是' if r['超限'] else '否', cell_style),
            ])
        rack_table = Table(rack_rows, colWidths=[30 * mm, 18 * mm, 18 * mm, 24 * mm, 22 * mm, 18 * mm, 16 * mm],
                           repeatRows=1)
        rack_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0ea5e9')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f3f4f6')]),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
        ]))
        story.append(rack_table)
    else:
        story.append(Paragraph('无机柜数据', normal_style))
    story.append(Spacer(1, 4 * mm))

    # 9. 校验结果
    story.append(Paragraph('9. 校验结果', h2_style))
    val = data['validation']
    status = '通过' if val.get('valid') else '失败'
    story.append(Paragraph(f"校验状态：<b>{status}</b>", normal_style))
    errors = val.get('errors', [])
    if errors:
        story.append(Paragraph('问题列表：', normal_style))
        for e in errors:
            story.append(Paragraph(f"  • {e}", normal_style))
    else:
        story.append(Paragraph('无校验问题，拓扑结构完整。', normal_style))

    # V2.7.4-T8: 构建时使用页眉页脚回调
    doc.build(story, onFirstPage=_draw_header_footer, onLaterPages=_draw_header_footer)
    return filename


def _generate_power_chart(power_data, font_name='Helvetica'):
    """V2.7.4-T7: 生成功率分布柱状图 PNG，返回临时文件路径"""
    try:
        import matplotlib
        matplotlib.use('Agg')
        import matplotlib.pyplot as plt
    except ImportError:
        return None

    # 提取功率数据（过滤非数值项）
    labels = []
    values = []
    for k, v in power_data.items():
        if isinstance(v, (int, float)) and v > 0:
            labels.append(str(k))
            values.append(float(v))

    if not values:
        return None

    import tempfile, os as _os
    fig, ax = plt.subplots(figsize=(7, 4))
    bars = ax.bar(labels, values, color=['#3b82f6', '#10b981', '#f59e0b', '#ef4444', '#8b5cf6'][:len(values)])
    ax.set_ylabel('功率 (W)', fontsize=10)
    ax.set_title('功率分布', fontsize=12)
    ax.tick_params(axis='x', rotation=30, labelsize=8)
    for bar, val in zip(bars, values):
        ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height(), f'{val:.0f}',
                ha='center', va='bottom', fontsize=8)
    fig.tight_layout()

    tmp_path = _os.path.join(tempfile.gettempdir(), 'autolink_power_chart.png')
    fig.savefig(tmp_path, dpi=150)
    plt.close(fig)
    return tmp_path


def _generate_module_cost_chart(modules_data, font_name='Helvetica'):
    """V2.7.4-T7: 生成光模块成本占比饼图 PNG，返回临时文件路径"""
    try:
        import matplotlib
        matplotlib.use('Agg')
        import matplotlib.pyplot as plt
    except ImportError:
        return None

    from optical_selector import estimate_module_cost

    labels = []
    costs = []
    for mid, info in modules_data.items():
        count = info.get('count', 0)
        price_range = info.get('price', '')
        if count > 0 and price_range:
            lo, hi = estimate_module_cost(price_range)
            avg_cost = (lo + hi) / 2 * count
            if avg_cost > 0:
                labels.append(str(mid))
                costs.append(avg_cost)

    if not costs:
        return None

    import tempfile, os as _os
    fig, ax = plt.subplots(figsize=(6, 4))
    ax.pie(costs, labels=labels, autopct='%1.1f%%', startangle=90,
           colors=['#3b82f6', '#10b981', '#f59e0b', '#ef4444', '#8b5cf6', '#ec4899'][:len(costs)])
    ax.set_title('光模块成本占比', fontsize=12)
    fig.tight_layout()

    tmp_path = _os.path.join(tempfile.gettempdir(), 'autolink_module_cost_chart.png')
    fig.savefig(tmp_path, dpi=150)
    plt.close(fig)
    return tmp_path
