"""AIDC 规划器 plan:table 契约 v1.2 测试（G0 桥接标识 + 契约 v1.2 身份/版本 + camelCase + 补齐字段）。

覆盖 backend/aidc_planner.py：
  - plan_aidc 输出符合契约 v1.2（meta 桥接标识 / projectId / planHash / planVersion / macro camelCase）
  - deviceList 逐设备含 rack
  - 规模映射（64 台 = 22 设备）与非法档位报错
  - validate_macro 兼容 camelCase/snake_case 输入
  - 导出 json/excel/zip（交付包）
"""
import io
import json
import zipfile

import pytest

from aidc_planner import (CONTRACT_VERSION, SCN_ABBR, _SCALE, export_plan,
                          import_plan, plan_aidc, plan_hash, validate_macro)


class TestPlanAidcContract:
    def test_bridge_meta(self):
        plan = plan_aidc({'gpu_count': 64})
        meta = plan['meta']
        assert meta['source'] == 'autolink'
        assert meta['projectType'] == 'aidc'
        assert meta['bridgeVersion'] == '1.0'
        assert meta['schema'] == 'plan:table/1.2'
        assert meta['version'] == CONTRACT_VERSION == '1.2'
        assert meta['generatedAt']
        assert meta['project'] == 'aidc_64'
        # 契约 v1.2：身份 + 版本必带
        assert meta['projectId']
        assert meta['planHash']
        assert meta['planVersion'] == 1

    def test_identity_project_id(self):
        pid = '7c9e6679-7425-40de-944b-e07fc1f90ae7'
        plan = plan_aidc({'gpu_count': 64, 'project_id': pid,
                          'project_name': 'H3C-64台-BJ01', 'plan_version': 3})
        meta = plan['meta']
        assert meta['projectId'] == pid
        assert meta['projectName'] == 'H3C-64台-BJ01'
        assert meta['planVersion'] == 3
        # projectId/projectName 不泄漏进 macro（planHash 只覆盖 macro）
        assert 'project_id' not in plan['macro'] and 'projectId' not in plan['macro']

    def test_project_id_minted_when_absent(self):
        a = plan_aidc({'gpu_count': 64})
        b = plan_aidc({'gpu_count': 64})
        assert a['meta']['projectId'] and b['meta']['projectId']
        assert a['meta']['projectId'] != b['meta']['projectId']  # 无 project_id 时每次 mint 新 UUID

    def test_plan_hash_deterministic_and_change_detection(self):
        a = plan_aidc({'gpu_count': 64})
        b = plan_aidc({'gpu_count': 64})
        assert a['meta']['planHash'] == b['meta']['planHash']  # 同 macro → 同 hash
        c = plan_aidc({'gpu_count': 64, 'pfc_queue': 4})
        assert a['meta']['planHash'] != c['meta']['planHash']  # 参数变 → hash 变
        assert c['meta']['planVersion'] == 1  # 未显式传 plan_version 时恒为 1（自增由项目层负责）

    def test_plan_hash_matches_util(self):
        plan = plan_aidc({'gpu_count': 64, 'project_id': 'x'})
        assert plan['meta']['planHash'] == plan_hash(plan['macro'])  # 与 canonical 算法一致
        # generatedAt 不参与哈希（meta 不在 macro 内）
        macro1 = {'site': 'BJ01', 'gpuCount': 64}
        macro2 = {'gpuCount': 64, 'site': 'BJ01'}
        assert plan_hash(macro1) == plan_hash(macro2)  # sort_keys 无关键序

    def test_macro_camelcase(self):
        plan = plan_aidc({'gpu_count': 64})
        m = plan['macro']
        assert m['gpuCount'] == 64
        assert m['pfcQueue'] == 3 and m['cnpQueue'] == 6 and m['bgpMaxPaths'] == 16
        assert m['naming']['abbr'] == SCN_ABBR
        assert set(m['ipSegments']) == {'loopback', 'compute', 'storage', 'biz', 'oob', 'interconnect'}
        assert m['ospf'] == {'process': 10, 'area': '0.0.0.0'}
        assert m['asRange'] == [65001, 65500]
        assert 'deviceModels' in m and 'vlanRanges' in m
        # 无 snake_case 残留
        assert 'pfc_queue' not in m and 'gpu_count' not in m

    def test_topology_and_protocols(self):
        plan = plan_aidc({'gpu_count': 64})
        assert plan['topology']['layers'] == 2
        assert plan['topology']['spines'] == 2 and plan['topology']['leaves'] == 8
        assert plan['protocols']['ospf']['process'] == 10
        assert plan['protocols']['bgp']['ecmp'] == 16

    def test_protocol_default_and_override(self):
        """49-a：macro.protocol 默认 RoCE；传入 IB 时自描述协议（示例库 IB/RoCE 区分）。"""
        plan = plan_aidc({'gpu_count': 64})
        assert plan['macro']['protocol'] == 'RoCE'
        ib = plan_aidc({'gpu_count': 64, 'protocol': 'IB'})
        assert ib['macro']['protocol'] == 'IB'
        assert ib['macro']['gpuCount'] == 64

    def test_protocol_affects_plan_hash(self):
        """protocol 是 macro 组成部分 → 不同协议 planHash 不同（变更检测权威判据）。"""
        roce = plan_aidc({'gpu_count': 64, 'protocol': 'RoCE'})
        ib = plan_aidc({'gpu_count': 64, 'protocol': 'IB'})
        assert roce['meta']['planHash'] != ib['meta']['planHash']

    def test_validate_macro_protocol(self):
        assert validate_macro({'gpu_count': 64, 'protocol': 'IB'}) is None
        assert validate_macro({'gpu_count': 64, 'protocol': 'RoCE'}) is None
        assert validate_macro({'gpu_count': 64, 'protocol': 'UEC'}) is None
        err = validate_macro({'gpu_count': 64, 'protocol': 'FCoE'})
        assert err and 'IB/RoCE/UEC' in err

    def test_device_list_has_rack(self):
        plan = plan_aidc({'gpu_count': 64})
        assert len(plan['deviceList']) == 22
        for d in plan['deviceList']:
            assert d['name'] and 'rack' in d
            assert d['rack'] == int(d['name'].split('-R', 1)[1].split('-', 1)[0])

    def test_scale_mapping(self):
        for gpu, (spine, leaf) in ((32, (2, 4)), (64, (2, 8)), (1024, (8, 32))):
            plan = plan_aidc({'gpu_count': gpu})
            assert plan['topology']['spines'] == spine
            assert plan['topology']['leaves'] == leaf

    def test_unsupported_scale_errors(self):
        plan = plan_aidc({'gpu_count': 96})
        assert 'error' in plan
        assert '不在支持档位' in plan['error']

    def test_validate_macro_accepts_camelcase(self):
        assert validate_macro({'gpuCount': 64}) is None
        assert validate_macro({'gpuCount': 96}) is not None

    def test_output_serializable(self):
        plan = plan_aidc({'gpu_count': 64})
        json.dumps(plan)  # 可序列化（供 CLI/GUI 输出）


class TestExport:
    """REQ-A3（G2）：plan:table 导出 json/excel + design:from-gpus 后端 action。"""

    def test_export_json(self, tmp_path):
        path = export_plan({'gpu_count': 64}, str(tmp_path / 'p.json'), 'json')
        data = json.load(io.open(path, encoding='utf-8'))
        assert data['meta']['projectType'] == 'aidc'
        assert len(data['deviceList']) == 22

    def test_export_excel(self, tmp_path):
        from openpyxl import load_workbook
        path = export_plan({'gpu_count': 64}, str(tmp_path / 'p'), 'excel')
        assert path.endswith('.xlsx')
        wb = load_workbook(path)
        for sheet in ('设备清单', '接线', '终端', '宏观参数', '协议', '收敛比'):
            assert sheet in wb.sheetnames

    def test_export_zip_delivery_package(self, tmp_path):
        """契约 v1.2（A-2）：交付包 ZIP 含 plan.json + README（版本戳）。"""
        path = export_plan({'gpu_count': 64, 'project_id': 'PID001', 'project_name': 'Demo'},
                           str(tmp_path / 'pkg'), 'zip')
        assert path.endswith('.zip')
        with zipfile.ZipFile(path) as zf:
            assert 'plan.json' in zf.namelist() and 'README.md' in zf.namelist()
            readme = zf.read('README.md').decode('utf-8')
            assert 'projectId' in readme and 'planVersion' in readme and 'planHash' in readme
            plan = json.loads(zf.read('plan.json').decode('utf-8'))
            assert plan['meta']['projectId'] == 'PID001'

    def test_export_zip_with_topology_png(self, tmp_path):
        """打磨轮（AL-B3）：交付包 ZIP 附带拓扑 PNG。"""
        import base64
        png = base64.b64encode(b'FAKEPNG').decode('utf-8')
        path = export_plan({'gpu_count': 64, 'project_id': 'PID', 'project_name': 'Demo'},
                           str(tmp_path / 'pkg'), 'zip', png_base64=png)
        with zipfile.ZipFile(path) as zf:
            names = zf.namelist()
            assert '拓扑图.png' in names
            assert zf.read('拓扑图.png') == b'FAKEPNG'

    def test_export_invalid_scale_raises(self):
        with pytest.raises(ValueError):
            export_plan({'gpu_count': 96}, 'x.json', 'json')

    def test_design_from_gpus_action(self):
        from cli import execute
        out = execute('design:from-gpus', {'gpu_count': 64})
        assert out['meta']['projectType'] == 'aidc'
        assert len(out['deviceList']) == 22


class TestImportPlan:
    """48-b（F8-2）：plan:table 回导（导入导出格式增强——JSON/ZIP 可回导）。"""

    def _plan(self, project_id='PID-IMPORT', plan_version=3, with_hash=True):
        plan = plan_aidc({'gpu_count': 64, 'project_id': project_id,
                          'project_name': '导入示例', 'plan_version': plan_version})
        meta = plan['meta']
        if not with_hash:
            meta.pop('planHash', None)
        return plan

    def test_import_plan_roundtrip(self):
        plan = self._plan()
        res = import_plan(plan)
        assert res['ok'] is True
        assert res['projectId'] == 'PID-IMPORT'
        assert res['planVersion'] == 3
        assert res['planHash'] == plan['meta']['planHash']
        assert res['plan']['meta']['projectId'] == 'PID-IMPORT'

    def test_import_plan_accepts_wrapper(self):
        """兼容 plan:aidc:export json 的外壳 {plan: {...}}（等价 zip 内 plan.json）。"""
        plan = self._plan()
        res = import_plan({'plan': plan})
        assert res['ok'] is True
        assert res['projectId'] == 'PID-IMPORT'

    def test_import_plan_recomputes_hash_when_missing(self):
        plan = self._plan(with_hash=False)
        assert 'planHash' not in plan['meta']
        res = import_plan(plan)
        assert res['ok'] is True
        # 归一化：按 macro 重算 planHash（契约 v1.2 §1.3）
        assert res['planHash'] == plan_hash(plan['macro'])
        assert plan['meta']['planHash'] == res['planHash']

    def test_import_plan_rejects_non_plan(self):
        assert import_plan({'schema': 'other', 'macro': {}, 'deviceList': []})['error']
        assert import_plan({'meta': {}})['error']
        assert import_plan('x')['error']
        assert import_plan(42)['error']

    def test_import_plan_rejects_incomplete_structure(self):
        plan = self._plan()
        del plan['terminals']
        res = import_plan(plan)
        assert 'error' in res and 'terminals' in res['error']

    def test_plan_aidc_import_action(self):
        """后端 action 层开放（engine.py handle_plan_aidc_import）。"""
        from cli import execute
        plan = self._plan()
        out = execute('plan:aidc:import', {'plan': plan})
        assert out['ok'] is True
        assert out['projectId'] == 'PID-IMPORT'
        assert out['plan']['meta']['planHash'] == plan['meta']['planHash']
