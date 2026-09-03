"""4.5 D-2 导出数据核对测试（F5-2：渲染批次产物 vs 设计/规划漂移）+ 5.0.1-501-d 导出内容级校验"""
import json

import pytest

from validation_engine import check_export_batch, collect_batch_stats, check_export_content


def _write_xlsx(path, rows):
    """写入一个 xlsx：首行为表头 + rows 条数据行"""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(['A端设备', 'A端接口', 'Z端设备'])
    for r in rows:
        ws.append(list(r))
    wb.save(path)
    return path


def _write_content_valid_xlsx(path, rows, sheets=('网络设计摘要', '服务器连接表', '参数网络连接表')):
    """写入连接表（多 sheet，5.0.1-501-d E009 契约；摘要 sheet 仅表头，与真实导出一致）"""
    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)
    wrote_rows = False
    for name in sheets:
        ws = wb.create_sheet(name)
        if name == '网络设计摘要':
            ws.append(['项目', '值'])
            continue  # 摘要不含连接行（与 exporter.export_all_connections 一致）
        ws.append(['A端设备', 'A端接口', 'Z端设备'])
        if not wrote_rows:
            for r in rows:
                ws.append(list(r))
            wrote_rows = True  # 连接行集中第一个连接 sheet，保证总行数 = len(rows)
    wb.save(path)
    return path


_DEVICE_LIST_HEADERS = ['设备类型', '厂商', '型号', '数量', '单机功耗(W)', 'U位高度', '总功耗(W)', '总U位']


def _write_device_list_xlsx(path, group_rows=(), total_qty=None):
    """写入设备清单（5.0.1-501-d E009/E010 契约）"""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = '设备清单'
    ws.append(_DEVICE_LIST_HEADERS)
    for r in group_rows:
        ws.append(list(r))
    if total_qty is not None:
        ws.append(['合计', '', '', total_qty, '', '', 0, 0])
    wb.save(path)
    return path


_CABLING_HEADERS = ['网络类型', 'A端设备', 'A端端口', 'A端机柜', 'A端U位', 'Z端设备', 'Z端端口',
                    'Z端机柜', 'Z端U位', '速率', '线缆类型', '1分2扇出', '光模块型号', '封装',
                    '规格', '光纤类型', '支持距离(m)', '估算长度(m)', '价格区间', '估价低(元)',
                    '估价高(元)', '描述']


def _write_cabling_xlsx(path, rows):
    """写入布线指导表（5.0.1-501-d E009/E011 契约）"""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = '布线指导表'
    ws.append(_CABLING_HEADERS)
    for r in rows:
        ws.append(list(r))
    wb.save(path)
    return path


def _manifest(servers=64, param_leaves=8, param_spines=2, mode='full'):
    return {
        'schema_version': 1,
        'project': 'P1',
        'version': 1,
        'config_hash': 'abc',
        'downlink_mode': mode,
        'output_types': ['connections', 'deviceList', 'bom'],
        'results': [{'type': 'connections', 'status': 'success'}],
        'stats': {
            'servers': servers,
            'param_leaves': param_leaves,
            'param_spines': param_spines,
            'param_cores': 0,
            'storage_leaves': 0,
            'storage_spines': 0,
            'biz_access': 0,
            'biz_agg': 0,
            'oob_access': 0,
            'oob_agg': 0,
        },
    }


def _design(servers=64, nets=None, connections=None, mode='full'):
    nets = nets or {'param_leaves': 8, 'param_spines': 2}
    return {
        'servers': servers,
        'mode': mode,
        'network_devices': nets,
        'connections': connections or [],
    }


class TestCollectBatchStats:
    def test_nonexistent_dir(self, tmp_path):
        stats = collect_batch_stats(str(tmp_path / 'missing'))
        assert stats['exists'] is False

    def test_reads_manifest_and_mode(self, tmp_path):
        batch = tmp_path / 'v1_ts'
        batch.mkdir()
        (batch / 'manifest.json').write_text(
            json.dumps(_manifest(mode='full'), ensure_ascii=False), encoding='utf-8')
        _write_xlsx(batch / 'AI智算网络_full模式_1.xlsx', [('a', 'p1', 'b')] * 5)
        stats = collect_batch_stats(str(batch))
        assert stats['has_manifest'] is True
        assert stats['mode'] == 'full'
        assert stats['connection_rows'] == 5

    def test_mode_inferred_from_filename_without_manifest(self, tmp_path):
        batch = tmp_path / 'v1_ts'
        batch.mkdir()
        _write_xlsx(batch / '设备清单_custom模式_1.xlsx', [('gpu', 'h3c', 'x')] * 3)
        stats = collect_batch_stats(str(batch))
        assert stats['mode'] == 'custom'
        assert stats['device_list_rows'] == 3


class TestCheckExportBatch:
    def test_missing_dir_e001(self, tmp_path):
        issues = check_export_batch(str(tmp_path / 'missing'))
        assert [i.rule_id for i in issues] == ['E001']
        assert issues[0].severity == 'error'

    def test_missing_manifest_warns_e002(self, tmp_path):
        batch = tmp_path / 'v1_ts'
        batch.mkdir()
        _write_xlsx(batch / 'AI智算网络_full模式_1.xlsx', [('a', 'p1', 'b')] * 2)
        issues = check_export_batch(str(batch))
        assert 'E002' in {i.rule_id for i in issues}

    def test_mode_drift_e003(self, tmp_path):
        batch = tmp_path / 'v1_ts'
        batch.mkdir()
        (batch / 'manifest.json').write_text(
            json.dumps(_manifest(mode='custom')), encoding='utf-8')
        issues = check_export_batch(str(batch), _design(mode='full'))
        e003 = [i for i in issues if i.rule_id == 'E003']
        assert len(e003) == 1
        assert e003[0].severity == 'error'

    def test_manifest_stats_drift_e004(self, tmp_path):
        batch = tmp_path / 'v1_ts'
        batch.mkdir()
        (batch / 'manifest.json').write_text(
            json.dumps(_manifest(servers=64, param_leaves=8)), encoding='utf-8')
        issues = check_export_batch(str(batch), _design(servers=64, nets={'param_leaves': 16, 'param_spines': 2}))
        e004 = [i for i in issues if i.rule_id == 'E004']
        assert len(e004) == 1
        assert 'param_leaves' in e004[0].location
        assert e004[0].severity == 'error'

    def test_connection_rows_drift_e005(self, tmp_path):
        batch = tmp_path / 'v1_ts'
        batch.mkdir()
        (batch / 'manifest.json').write_text(
            json.dumps(_manifest()), encoding='utf-8')
        _write_xlsx(batch / 'AI智算网络_full模式_1.xlsx', [('a', 'p1', 'b')] * 3)
        design = _design(connections=[{'source': 'a', 'target': 'b'}] * 5)
        issues = check_export_batch(str(batch), design)
        e005 = [i for i in issues if i.rule_id == 'E005']
        assert len(e005) == 1
        assert e005[0].severity == 'error'

    def test_empty_device_list_e007(self, tmp_path):
        batch = tmp_path / 'v1_ts'
        batch.mkdir()
        (batch / 'manifest.json').write_text(
            json.dumps(_manifest()), encoding='utf-8')
        _write_xlsx(batch / '设备清单_full模式_1.xlsx', [])  # 仅表头
        issues = check_export_batch(str(batch), _design())
        e007 = [i for i in issues if i.rule_id == 'E007']
        assert len(e007) >= 1

    def test_filename_mode_mismatch_e006(self, tmp_path):
        batch = tmp_path / 'v1_ts'
        batch.mkdir()
        (batch / 'manifest.json').write_text(
            json.dumps(_manifest(mode='full')), encoding='utf-8')
        _write_xlsx(batch / 'AI智算网络_custom模式_1.xlsx', [('a', 'p1', 'b')])
        issues = check_export_batch(str(batch), _design(mode='full'))
        e006 = [i for i in issues if i.rule_id == 'E006']
        assert len(e006) == 1

    def test_clean_batch_no_error(self, tmp_path):
        batch = tmp_path / 'v1_ts'
        batch.mkdir()
        (batch / 'manifest.json').write_text(
            json.dumps(_manifest(servers=64, param_leaves=8, param_spines=2)), encoding='utf-8')
        # 5.0.1-501-d: 连接表/设备清单采用内容契约（多 sheet + 表头 + 合计数量）
        _write_content_valid_xlsx(batch / 'AI智算网络_full模式_1.xlsx', [('a', 'p1', 'b')] * 5)
        _write_device_list_xlsx(batch / '设备清单_full模式_1.xlsx',
                                group_rows=[('GPU服务器', 'NVIDIA', 'DGX-H100', 64, 2000, 4, 0, 0)],
                                total_qty=74)
        design = _design(servers=64, nets={'param_leaves': 8, 'param_spines': 2},
                         connections=[{'source': 'a', 'target': 'b'}] * 5)
        issues = check_export_batch(str(batch), design)
        assert not [i for i in issues if i.severity == 'error']
        assert not [i for i in issues if i.rule_id in ('E009', 'E010', 'E011')]


class TestBatchIntegrity:
    """48-e（F8-5）：批次产物完整性——manifest.files 逐文件（缺失/漂移/哈希不符 → E008）"""

    def _sha256(self, path):
        import hashlib
        return hashlib.sha256(path.read_bytes()).hexdigest()

    def _write_manifest_with_files(self, batch, files, extra_actual=()):
        """files: [{name, content}] → 写文件并生成带逐文件清单的 manifest"""
        for f in files:
            (batch / f['name']).write_bytes(f['content'])
        for name in extra_actual:
            (batch / name).write_text('drift', encoding='utf-8')
        (batch / 'manifest.json').write_text(json.dumps({
            'schema_version': 1, 'version': 1, 'config_hash': 'x',
            'output_types': ['connections'], 'results': [],
            'stats': {'servers': 0, 'param_leaves': 0, 'param_spines': 0, 'param_cores': 0,
                      'storage_leaves': 0, 'storage_spines': 0, 'biz_access': 0, 'biz_agg': 0,
                      'oob_access': 0, 'oob_agg': 0},
            'files': [{'name': f['name'], 'size': len(f['content']), 'sha256': self._sha256(batch / f['name'])}
                      for f in files],
        }, ensure_ascii=False), encoding='utf-8')

    def test_intact_batch_no_e008(self, tmp_path):
        batch = tmp_path / 'v1_ts'
        batch.mkdir()
        self._write_manifest_with_files(batch, [{'name': 'a.xlsx', 'content': b'AAA'}])
        issues = check_export_batch(str(batch))
        assert not [i for i in issues if i.rule_id == 'E008']

    def test_missing_file_e008(self, tmp_path):
        batch = tmp_path / 'v1_ts'
        batch.mkdir()
        files = [{'name': 'a.xlsx', 'content': b'AAA'}, {'name': 'b.xlsx', 'content': b'BBB'}]
        self._write_manifest_with_files(batch, files)
        (batch / 'b.xlsx').unlink()
        issues = check_export_batch(str(batch))
        e008 = [i for i in issues if i.rule_id == 'E008']
        assert any('缺失' in i.message for i in e008)

    def test_hash_mismatch_e008(self, tmp_path):
        batch = tmp_path / 'v1_ts'
        batch.mkdir()
        self._write_manifest_with_files(batch, [{'name': 'a.xlsx', 'content': b'AAA'}])
        # 篡改文件内容 → 哈希不符
        (batch / 'a.xlsx').write_bytes(b'BBB')
        issues = check_export_batch(str(batch))
        e008 = [i for i in issues if i.rule_id == 'E008']
        assert any('哈希不符' in i.message for i in e008)

    def test_drift_extra_file_e008(self, tmp_path):
        batch = tmp_path / 'v1_ts'
        batch.mkdir()
        self._write_manifest_with_files(batch, [{'name': 'a.xlsx', 'content': b'AAA'}], extra_actual=('extra.txt',))
        issues = check_export_batch(str(batch))
        e008 = [i for i in issues if i.rule_id == 'E008']
        assert any('漂移' in i.message for i in e008)

    def test_check_batch_integrity_returns_problems(self, tmp_path):
        """独立入口：check_batch_integrity 直接校验。"""
        from validation_engine.export_check import check_batch_integrity
        batch = tmp_path / 'v1_ts'
        batch.mkdir()
        self._write_manifest_with_files(batch, [{'name': 'a.xlsx', 'content': b'AAA'}])
        (batch / 'a.xlsx').write_bytes(b'ZZZ')
        problems = check_batch_integrity(str(batch))
        assert any(p.rule_id == 'E008' and '哈希不符' in p.message for p in problems)
        assert all(p.rule_id == 'E008' for p in problems)


class TestExportContent:
    """5.0.1-501-d: 导出内容级校验（E009 表头契约 / E010 关键值 / E011 行数一致）"""

    _DESIGN = {
        'servers': 64,
        'network_devices': {'param_leaves': 8, 'param_spines': 2},
        'connections': [{'source': 'a', 'target': 'b'}] * 5,
    }

    def test_clean_content_no_e009_e010_e011(self, tmp_path):
        batch = tmp_path / 'v1_ts'
        batch.mkdir()
        _write_content_valid_xlsx(batch / 'AI智算网络_full模式_1.xlsx', [('a', 'p1', 'b')] * 5)
        _write_device_list_xlsx(batch / '设备清单_full模式_1.xlsx', total_qty=74)
        _write_cabling_xlsx(batch / '布线指导表_full模式_1.xlsx', [('param', 'a', 'p1', 'b', 'p2')] * 5)
        problems = check_export_content(str(batch), self._DESIGN)
        assert not [p for p in problems if p.rule_id in ('E009', 'E010', 'E011')]

    def test_device_list_header_drift_e009(self, tmp_path):
        batch = tmp_path / 'v1_ts'
        batch.mkdir()
        _write_content_valid_xlsx(batch / 'AI智算网络_full模式_1.xlsx', [])
        _write_xlsx(batch / '设备清单_full模式_1.xlsx', [('gpu', 'h3c', 'x')])
        problems = check_export_content(str(batch), self._DESIGN)
        e009 = [p for p in problems if p.rule_id == 'E009']
        assert any('表头' in p.message for p in e009)

    def test_missing_device_list_sheet_e009(self, tmp_path):
        batch = tmp_path / 'v1_ts'
        batch.mkdir()
        from openpyxl import Workbook
        wb = Workbook()
        wb.active.append(['设备类型', '厂商', '型号', '数量'])
        wb.save(batch / '设备清单_full模式_1.xlsx')
        problems = check_export_content(str(batch), self._DESIGN)
        e009 = [p for p in problems if p.rule_id == 'E009']
        assert any('缺少 sheet' in p.message or '表头' in p.message for p in e009)

    def test_missing_connections_sheet_e009(self, tmp_path):
        batch = tmp_path / 'v1_ts'
        batch.mkdir()
        _write_xlsx(batch / 'AI智算网络_full模式_1.xlsx', [('a', 'p1', 'b')])
        problems = check_export_content(str(batch), self._DESIGN)
        e009 = [p for p in problems if p.rule_id == 'E009']
        assert any('关键 sheet' in p.message for p in e009)

    def test_device_list_total_mismatch_e010(self, tmp_path):
        batch = tmp_path / 'v1_ts'
        batch.mkdir()
        _write_device_list_xlsx(batch / '设备清单_full模式_1.xlsx', total_qty=60)
        problems = check_export_content(str(batch), self._DESIGN)
        e010 = [p for p in problems if p.rule_id == 'E010']
        assert len(e010) == 1
        assert '60' in e010[0].message and '74' in e010[0].message

    def test_cabling_rows_mismatch_e011(self, tmp_path):
        batch = tmp_path / 'v1_ts'
        batch.mkdir()
        _write_cabling_xlsx(batch / '布线指导表_full模式_1.xlsx', [('param', 'a', 'p1', 'b', 'p2')] * 3)
        problems = check_export_content(str(batch), self._DESIGN)
        e011 = [p for p in problems if p.rule_id == 'E011']
        assert len(e011) == 1
        assert '3' in e011[0].message and '5' in e011[0].message

    def test_cabling_missing_sheet_e009(self, tmp_path):
        batch = tmp_path / 'v1_ts'
        batch.mkdir()
        from openpyxl import Workbook
        wb = Workbook()
        wb.active.append(['x'])
        wb.save(batch / '布线指导表_full模式_1.xlsx')
        problems = check_export_content(str(batch), self._DESIGN)
        e009 = [p for p in problems if p.rule_id == 'E009']
        assert any('布线指导表' in p.message or '表头' in p.message for p in e009)

    def test_check_export_batch_runs_content_checks(self, tmp_path):
        """check_export_batch 自动聚合 E009/E010/E011（5.0.1-501-d 接入门禁）"""
        batch = tmp_path / 'v1_ts'
        batch.mkdir()
        (batch / 'manifest.json').write_text(json.dumps(_manifest()), encoding='utf-8')
        _write_xlsx(batch / 'AI智算网络_full模式_1.xlsx', [('a', 'p1', 'b')])  # 缺关键 sheet → E009
        _write_device_list_xlsx(batch / '设备清单_full模式_1.xlsx', total_qty=60)  # 合计不符 → E010
        issues = check_export_batch(str(batch), self._DESIGN)
        assert 'E009' in {i.rule_id for i in issues}
        assert 'E010' in {i.rule_id for i in issues}
