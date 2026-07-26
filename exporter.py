"""
AutoLink V1.1 - Excel导出与格式化
提供服务器视角、交换机视角的连接表生成和Excel美化
"""

import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.utils import get_column_letter


def _extract_number(s):
    """从字符串中提取数字用于排序"""
    if isinstance(s, (int, float)):
        return int(s)
    if not isinstance(s, str):
        return 0
    match = re.search(r'\d+', s)
    return int(match.group()) if match else 0


def _get_iface_weight(port_str):
    """获取服务器接口类型排序权重: 参数网卡=1, 存储=2, OOB=3, 业务=4"""
    if '参数' in port_str:
        return 1
    if '存储' in port_str:
        return 2
    if 'OOB' in port_str:
        return 3
    if '业务' in port_str:
        return 4
    return 5


def _get_switch_type_weight(device_name):
    """获取交换机类型权重用于排序: Leaf=1, Spine=2, Core=3, 接入=4, 汇聚=5"""
    if "Leaf" in device_name:
        return 1
    elif "Spine" in device_name:
        return 2
    elif "Core" in device_name:
        return 3
    elif "接入" in device_name:
        return 4
    elif "汇聚" in device_name:
        return 5
    return 6


def generate_server_view(designer):
    """生成服务器视角的连接表"""
    connections = []

    for server in designer.servers:
        server_group = designer.server_groups[server.name]
        podid = designer.podid_map.get(server.name, "")
        for conn in server.connections:
            if conn.a_device == server.name:
                connections.append({
                    'podid': podid,
                    '服务器分组': server_group,
                    'A端设备': conn.a_device,
                    'A端接口': conn.a_port,
                    'A端模块': conn.a_module,
                    'Z端设备': conn.z_device,
                    'Z端接口': conn.z_port,
                    'Z端模块': conn.z_module,
                    '线缆类型': conn.cable_type,
                    '描述': conn.description
                })

    df = pd.DataFrame(connections)
    if not df.empty:
        # 按设备聚合, 同设备内按接口类型排序: 参数→存储→OOB→业务
        df['iface_w'] = df['A端接口'].apply(_get_iface_weight)
        df['port_num'] = df['A端接口'].apply(_extract_number)
        # 从设备名末尾提取数字用于自然排序 (GPU服务器_2 < GPU服务器_10)
        df['dev_num'] = df['A端设备'].apply(_extract_number)
        df = df.sort_values(by=['服务器分组', 'dev_num', 'iface_w', 'port_num'])
        df = df.drop(columns=['dev_num', 'iface_w', 'port_num'])
    return df


def generate_switch_view(designer):
    """生成交换机视角的连接表"""
    param_connections = []
    storage_connections = []

    all_switches = (designer.param_leaves + designer.param_spines + designer.param_cores +
                   designer.storage_leaves + designer.storage_spines + designer.storage_cores)

    for switch in all_switches:
        switch_group = designer.switch_groups.get(switch.name, "")
        podid = designer.podid_map.get(switch.name, "")
        for conn in switch.connections:
            if conn.a_device == switch.name:
                row = {
                    'podid': podid,
                    '交换机分组': switch_group,
                    'A端设备': conn.a_device,
                    'A端接口': conn.a_port,
                    'A端模块': conn.a_module,
                    'Z端设备': conn.z_device,
                    'Z端接口': conn.z_port,
                    'Z端模块': conn.z_module,
                    '线缆类型': conn.cable_type,
                    '描述': conn.description
                }
                if "参数" in switch.name:
                    param_connections.append(row)
                else:
                    storage_connections.append(row)

    param_df = pd.DataFrame(param_connections)
    storage_df = pd.DataFrame(storage_connections)

    if not param_df.empty:
        param_df['type_weight'] = param_df['A端设备'].apply(_get_switch_type_weight)
        param_df['dev_num'] = param_df['A端设备'].apply(_extract_number)
        param_df['port_num'] = param_df['A端接口'].apply(_extract_number)
        param_df = param_df.sort_values(by=['type_weight', '交换机分组', 'dev_num', 'A端设备', 'port_num'])
        param_df = param_df.drop(columns=['type_weight', 'dev_num', 'port_num'])

    if not storage_df.empty:
        storage_df['type_weight'] = storage_df['A端设备'].apply(_get_switch_type_weight)
        storage_df['dev_num'] = storage_df['A端设备'].apply(_extract_number)
        storage_df['port_num'] = storage_df['A端接口'].apply(_extract_number)
        storage_df = storage_df.sort_values(by=['type_weight', '交换机分组', 'dev_num', 'A端设备', 'port_num'])
        storage_df = storage_df.drop(columns=['type_weight', 'dev_num', 'port_num'])

    return {'参数网络': param_df, '存储网络': storage_df}


def generate_summary_data(designer):
    """生成网络设计摘要数据"""
    half_ports = designer.param_switch_ports // 2
    if designer.param_3tier_needed:
        servers_per_group = min(designer.param_servers_per_pod // designer.param_ports_per_server, half_ports)
        param_downlink_usage = f"{servers_per_group}/{half_ports} ({servers_per_group / half_ports * 100:.1f}%)"
        param_uplink_usage = f"{half_ports}/{half_ports} (100.0%)"
    else:
        param_downlink_usage = f"{designer.param_servers_per_group}/{half_ports} ({designer.param_servers_per_group / half_ports * 100:.1f}%)"
        uplinks_used = min(half_ports, designer.param_spine_count)
        param_uplink_usage = f"{uplinks_used}/{half_ports} ({uplinks_used / half_ports * 100:.1f}%)"

    half_ports = designer.storage_switch_ports // 2
    if designer.storage_3tier_needed:
        servers_per_group = min(designer.storage_servers_per_pod // designer.storage_ports_per_server, half_ports)
        storage_downlink_usage = f"{servers_per_group}/{half_ports} ({servers_per_group / half_ports * 100:.1f}%)"
        storage_uplink_usage = f"{half_ports}/{half_ports} (100.0%)"
    else:
        storage_downlink_usage = f"{designer.storage_servers_per_group}/{half_ports} ({designer.storage_servers_per_group / half_ports * 100:.1f}%)"
        uplinks_used = min(half_ports, designer.storage_spine_count)
        storage_uplink_usage = f"{uplinks_used}/{half_ports} ({uplinks_used / half_ports * 100:.1f}%)"

    param_tier_info = "3层(Leaf-Spine-Core)" if designer.param_3tier_needed else "2层(Leaf-Spine)"
    storage_tier_info = "3层(Leaf-Spine-Core)" if designer.storage_3tier_needed else "2层(Leaf-Spine)"

    if designer.param_3tier_needed:
        param_group_info = f"{designer.param_pods}个POD, 每个POD{designer.param_servers_per_pod}台"
    else:
        param_group_info = f"{designer.param_groups}组, 每组{designer.param_servers_per_group}台"

    if designer.storage_3tier_needed:
        storage_group_info = f"{designer.storage_pods}个POD, 每个POD{designer.storage_servers_per_pod}台"
    else:
        storage_group_info = f"{designer.storage_groups}组, 每组{designer.storage_servers_per_group}台"

    summary = [
        ["网络设计摘要", ""],
        ["设计时间", pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["", ""],
        ["GPU服务器配置", ""],
        ["服务器数量", designer.num_servers],
        ["参数网卡/服务器", designer.param_ports_per_server],
        ["存储网卡/服务器", designer.storage_ports_per_server],
        ["", ""],
        ["参数网络设计", ""],
        ["交换机端口数", designer.param_switch_ports],
        ["Leaf交换机数量", designer.param_leaf_count],
        ["Spine交换机数量", designer.param_spine_count],
        ["Core交换机数量", designer.param_core_count if designer.param_3tier_needed else "无"],
        ["网络层级", param_tier_info],
        ["服务器分组", param_group_info],
        ["下行端口使用率", param_downlink_usage],
        ["上行端口使用率", param_uplink_usage],
        ["收敛比例", "1:1:1" if designer.param_3tier_needed else "1:1"],
        ["", ""],
        ["存储网络设计", ""],
        ["交换机端口数", designer.storage_switch_ports],
        ["Leaf交换机数量", designer.storage_leaf_count],
        ["Spine交换机数量", designer.storage_spine_count],
        ["Core交换机数量", designer.storage_core_count if designer.storage_3tier_needed else "无"],
        ["网络层级", storage_tier_info],
        ["服务器分组", storage_group_info],
        ["下行端口使用率", storage_downlink_usage],
        ["上行端口使用率", storage_uplink_usage],
        ["收敛比例", "1:1:1" if designer.storage_3tier_needed else "1:1"],
        ["", ""],
        ["网络速度配置", ""],
        ["参数网络速度", designer.param_speed],
        ["存储网络速度", designer.storage_speed],
        ["线缆类型", "参数: MPO | 存储: AOC | OOB: 网线/光纤 | 业务: 光纤"]
    ]
    return pd.DataFrame(summary, columns=["项目", "值"])


def apply_excel_formatting(filename):
    """应用Excel格式美化：合并同设备行、组分隔线、边框、交替行色"""
    from openpyxl.styles import Border, Side, PatternFill, Font, Alignment

    wb = load_workbook(filename)
    thin = Side(style='thin')
    thick = Side(style='medium')
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    thick_bottom = Border(left=thin, right=thin, top=thin, bottom=thick)
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(bold=True, size=12, color="FFFFFF")
    group_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    even_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    odd_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    col_widths = {
        'podid': 15, '服务器分组': 15, '交换机分组': 15,
        'A端设备': 30, 'A端接口': 15, 'A端模块': 12,
        'Z端设备': 30, 'Z端接口': 15, 'Z端模块': 12,
        '线缆类型': 12, '描述': 35, '项目': 25, '值': 40
    }

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        max_row = ws.max_row
        max_col = ws.max_column

        if max_row <= 1:
            continue

        ws.freeze_panes = 'A3'
        ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"

        # 列宽
        for col_idx, col_name in enumerate(ws[1], 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_name.value, 15)

        ws.row_dimensions[1].height = 28

        # 标题行
        for c in ws[1]:
            c.font = header_font; c.fill = header_fill
            c.alignment = center_align; c.border = thin_border

        # 识别列：从标题行确定podid_col, group_col, device_col, port_col
        headers = {cell.value: cell.column for cell in ws[1] if cell.value}
        podid_col = headers.get('podid')
        group_col = headers.get('服务器分组') or headers.get('交换机分组')
        device_col = headers.get('A端设备')
        port_col = headers.get('A端接口')

        # 读取所有数据值（处理NaN → ffill）
        if max_row >= 2:
            rows_data = []
            for row_idx in range(2, max_row + 1):
                row_vals = {}
                for col in [podid_col, group_col, device_col, port_col]:
                    if col:
                        v = ws.cell(row=row_idx, column=col).value
                        row_vals[col] = v
                rows_data.append(row_vals)

            # Forward-fill NaN values (pandas output leaves merged cells as NaN)
            for i in range(len(rows_data)):
                if i > 0:
                    for col in [podid_col, group_col, device_col]:
                        if col and rows_data[i].get(col) is None:
                            rows_data[i][col] = rows_data[i-1].get(col)

            # 写回填充值
            for i, rd in enumerate(rows_data):
                r = i + 2
                for col, val in rd.items():
                    if col and ws.cell(row=r, column=col).value is None and val is not None:
                        ws.cell(row=r, column=col).value = val

        # 摘要sheet特殊处理
        if sheet_name in ('网络设计摘要', '设计摘要'):
            for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
                for cell in row:
                    cell.alignment = center_align
                    cell.border = thin_border
            continue

        # ===== 收集合并区域和组分界 =====
        merge_regions = []
        group_boundaries = set()

        # podid列合并
        if podid_col:
            cur, start = None, 2
            for r in range(2, max_row + 1):
                v = ws.cell(row=r, column=podid_col).value
                if v != cur:
                    if cur is not None and r - 1 > start:
                        merge_regions.append((start, r - 1, podid_col, podid_col))
                    cur, start = v, r
            if cur is not None and max_row > start:
                merge_regions.append((start, max_row, podid_col, podid_col))

        # 分组列合并 + 组分界标记
        if group_col:
            cur, start = None, 2
            for r in range(2, max_row + 1):
                v = ws.cell(row=r, column=group_col).value
                if v != cur:
                    if cur is not None:
                        if r - 1 > start:
                            merge_regions.append((start, r - 1, group_col, group_col))
                        group_boundaries.add(r - 1)  # 组结束行
                    cur, start = v, r
            if cur is not None and max_row > start:
                merge_regions.append((start, max_row, group_col, group_col))

        # 设备列合并（在每个分组内独立合并）
        if device_col and group_col:
            for r in range(2, max_row + 1):
                if r == 2:
                    cur_dev, dev_start = ws.cell(row=r, column=device_col).value, r
                    continue
                if r in group_boundaries:
                    continue  # 不在边界重置, 让下一行group变更触发合并
                dev_val = ws.cell(row=r, column=device_col).value
                grp_val = ws.cell(row=r, column=group_col).value
                prev_grp = ws.cell(row=r - 1, column=group_col).value
                if grp_val != prev_grp:
                    if r - 1 > dev_start:
                        merge_regions.append((dev_start, r - 1, device_col, device_col))
                    cur_dev, dev_start = dev_val, r
                elif dev_val != cur_dev:
                    if r - 1 > dev_start:
                        merge_regions.append((dev_start, r - 1, device_col, device_col))
                    cur_dev, dev_start = dev_val, r
            if max_row > dev_start:
                merge_regions.append((dev_start, max_row, device_col, device_col))
        elif device_col:
            cur, start = None, 2
            for r in range(2, max_row + 1):
                v = ws.cell(row=r, column=device_col).value
                if v != cur:
                    if cur is not None and r - 1 > start:
                        merge_regions.append((start, r - 1, device_col, device_col))
                    cur, start = v, r
            if cur is not None and max_row > start:
                merge_regions.append((start, max_row, device_col, device_col))

        # 应用合并
        for sr, er, sc, ec in merge_regions:
            ws.merge_cells(start_row=sr, end_row=er, start_column=sc, end_column=ec)

        # ===== 数据行样式 (单次遍历) =====
        for row_idx in range(2, max_row + 1):
            is_boundary = row_idx in group_boundaries
            fill = even_fill if row_idx % 2 == 0 else odd_fill
            row = [ws.cell(row=row_idx, column=c) for c in range(1, max_col + 1)]

            for cell in row:
                cell.border = thick_bottom if is_boundary else thin_border
                if is_boundary and cell.column == group_col:
                    cell.fill = group_fill
                else:
                    cell.fill = fill
                col = cell.column
                if col == podid_col or col == group_col or col == device_col:
                    cell.alignment = center_align
                elif col == max_col or (port_col and col > device_col):
                    cell.alignment = left_align if col == max_col else center_align
                else:
                    cell.alignment = center_align

        # 端口列居左
        if port_col:
            for r in range(2, max_row + 1):
                ws.cell(row=r, column=port_col).alignment = center_align

    wb.save(filename)
    print(f"已应用Excel格式美化: {filename}")


def export_all_connections(designer, filename):
    """导出所有连接关系到单个Excel文件"""
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        try:
            generate_summary_data(designer).to_excel(writer, sheet_name='网络设计摘要', index=False)
            generate_server_view(designer).to_excel(writer, sheet_name='服务器连接表', index=False)
            switch_views = generate_switch_view(designer)
            switch_views['参数网络'].to_excel(writer, sheet_name='参数网络连接表', index=False)
            switch_views['存储网络'].to_excel(writer, sheet_name='存储网络连接表', index=False)

            # OOB网络
            if designer.oob_enabled and designer.oob_access:
                oob_df = _generate_access_agg_view(designer, designer.oob_access, designer.oob_agg)
                oob_df.to_excel(writer, sheet_name='OOB网络连接表', index=False)
                # OOB汇聚反向视角
                oob_rev = _generate_agg_reverse_view(designer, designer.oob_access, designer.oob_agg, "OOB")
                if oob_rev is not None:
                    oob_rev.to_excel(writer, sheet_name='OOB汇聚视角', index=False)

            # 业务网络
            if designer.biz_enabled and designer.biz_access:
                biz_df = _generate_access_agg_view(designer, designer.biz_access, designer.biz_agg)
                biz_df.to_excel(writer, sheet_name='业务网络连接表', index=False)

        except Exception as e:
            print(f"生成连接表时出错: {e}")
            if len(writer.book.sheetnames) == 0:
                writer.book.create_sheet("错误信息")
            raise

    apply_excel_formatting(filename)
    print(f"所有连接表已导出到: {filename}")


def _generate_access_agg_view(designer, access_switches, agg_switches):
    """生成接入-汇聚网络交换机视角的连接表"""
    connections = []
    all_switches = access_switches + agg_switches

    for sw in all_switches:
        sw_group = designer.switch_groups.get(sw.name, "")
        podid = designer.podid_map.get(sw.name, "")
        for conn in sw.connections:
            if conn.a_device == sw.name:
                connections.append({
                    'podid': podid,
                    '交换机分组': sw_group,
                    'A端设备': conn.a_device,
                    'A端接口': conn.a_port,
                    'A端模块': conn.a_module,
                    'Z端设备': conn.z_device,
                    'Z端接口': conn.z_port,
                    'Z端模块': conn.z_module,
                    '线缆类型': conn.cable_type,
                    '描述': conn.description
                })

    df = pd.DataFrame(connections)
    if not df.empty:
        df['type_weight'] = df['A端设备'].apply(_get_switch_type_weight)
        df['dev_num'] = df['A端设备'].apply(_extract_number)
        df['port_num'] = df['A端接口'].apply(_extract_number)
        df = df.sort_values(by=['type_weight', '交换机分组', 'dev_num', 'A端设备', 'port_num'])
        df = df.drop(columns=['type_weight', 'dev_num', 'port_num'])
    return df


def _generate_agg_reverse_view(designer, access_switches, agg_switches, network_name):
    """生成汇聚交换机反向视角 (A端=汇聚, Z端=接入)"""
    connections = []

    for agg in agg_switches:
        sw_group = designer.switch_groups.get(agg.name, "")
        podid = designer.podid_map.get(agg.name, "")
        for conn in agg.connections:
            if conn.a_device == agg.name:
                connections.append({
                    'podid': podid,
                    '交换机分组': sw_group,
                    'A端设备': conn.a_device,
                    'A端接口': conn.a_port,
                    'A端模块': conn.a_module,
                    'Z端设备': conn.z_device,
                    'Z端接口': conn.z_port,
                    'Z端模块': conn.z_module,
                    '线缆类型': conn.cable_type,
                    '描述': conn.description
                })

    df = pd.DataFrame(connections)
    if not df.empty:
        df['dev_num'] = df['A端设备'].apply(_extract_number)
        df['port_num'] = df['A端接口'].apply(_extract_number)
        df = df.sort_values(by=['交换机分组', 'dev_num', 'A端设备', 'port_num'])
        df = df.drop(columns=['dev_num', 'port_num'])
    return df if not df.empty else None
