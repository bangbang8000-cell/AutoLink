"""
AutoLink V1.1 - 智能网络设计工具
修复问题：
1. (FIX-01) 修复三层Leaf数量计算错误：leafs_per_pod从固定ports_per_server改为动态计算groups_per_pod*ports_per_server
2. (FIX-02) 修复二层最大容量公式：k^2/(2p) → k^2/(4p)，符合标准Fat-Tree公式
3. (FIX-03) 修复三层Spine/Core连接：使用轮转(Round-Robin)分配防止端口溢出
4. (FIX-04) 修复存储网络：独立计算，不再依赖参数网络的POD结构
5. (FIX-05) 修复端口使用率计算：分别显示下联/上联端口使用率
6. (T1.5) 添加拓扑合法性自检：自动检查服务器覆盖率、端口溢出、收敛比
"""

import math
import re
import datetime
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.utils import get_column_letter
from collections import defaultdict
import configparser
import os


class NetworkObject:
    def __init__(self, name, obj_type, group=None, max_ports=64, podid=None):
        self.name = name
        self.obj_type = obj_type  # 'server', 'param_leaf', 'param_spine', 'param_core', 'storage_leaf', 'storage_spine'
        self.group = group
        self.podid = podid
        self.connections = []
        self.max_ports = max_ports

        # 根据设备类型初始化端口计数器
        if "leaf" in obj_type:
            # Leaf: 前一半端口用于服务器，后一半端口用于Spine
            self.downlink_counter = 1
            self.uplink_counter = math.ceil(max_ports / 2) + 1
            self.downlink_limit = math.floor(max_ports / 2)
            self.uplink_limit = max_ports
        elif "spine" in obj_type:
            # Spine: 前一半端口用于Leaf，后一半端口用于Core
            self.downlink_counter = 1
            self.uplink_counter = math.ceil(max_ports / 2) + 1
            self.downlink_limit = math.floor(max_ports / 2)
            self.uplink_limit = max_ports
        elif "core" in obj_type:
            # Core: 所有端口用于Spine连接
            self.core_counter = 1
            self.core_limit = max_ports
        else:
            # 服务器使用普通端口计数器
            self.port_counter = 1
            self.port_limit = max_ports

    def add_connection(self, connection):
        """添加连接关系"""
        self.connections.append(connection)

    def get_downlink_port(self):
        """获取下联端口(用于连接服务器或Leaf)"""
        if self.downlink_counter > self.downlink_limit:
            raise ValueError(f"{self.name}的下联端口数量超过限制({self.downlink_limit})")
        port_num = self.downlink_counter
        self.downlink_counter += 1
        return f"端口{port_num}"

    def get_uplink_port(self):
        """获取上联端口(用于连接Spine或Core)"""
        if self.uplink_counter > self.uplink_limit:
            raise ValueError(f"{self.name}的上联端口数量超过限制({self.uplink_limit - math.floor(self.max_ports / 2)})")
        port_num = self.uplink_counter
        self.uplink_counter += 1
        return f"端口{port_num}"

    def get_core_port(self):
        """获取Core交换机端口"""
        if self.core_counter > self.core_limit:
            raise ValueError(f"{self.name}的端口数量超过最大值{self.core_limit}")
        port_num = self.core_counter
        self.core_counter += 1
        return f"端口{port_num}"

    def get_server_port(self):
        """获取服务器端口"""
        if self.port_counter > self.port_limit:
            raise ValueError(f"{self.name}的端口数量超过最大值{self.port_limit}")
        port_num = self.port_counter
        self.port_counter += 1
        return f"端口{port_num}"

    def get_next_port(self, start=None):
        """兼容旧方法，获取下一个可用端口号"""
        if "leaf" in self.obj_type or "spine" in self.obj_type:
            # 尝试使用下联端口
            try:
                return self.get_downlink_port()
            except ValueError:
                return self.get_uplink_port()
        elif "core" in self.obj_type:
            return self.get_core_port()
        else:
            return self.get_server_port()


class Connection:
    def __init__(self, a_device, a_port, a_module, z_device, z_port, z_module, cable_type, description):
        self.a_device = a_device
        self.a_port = a_port
        self.a_module = a_module
        self.z_device = z_device
        self.z_port = z_port
        self.z_module = z_module
        self.cable_type = cable_type
        self.description = description


class ThreeTierNetwork:
    """三层胖树网络设计器"""

    def __init__(self, ports_per_server, switch_ports, network_speed, cable_type_config, network_type="param"):
        self.ports_per_server = ports_per_server
        self.switch_ports = switch_ports
        self.network_speed = network_speed
        self.cable_type_config = cable_type_config
        self.network_type = network_type
        self.prefix = "参数" if network_type == "param" else "存储"

        # 网络组件
        self.leaves = []
        self.spines = []
        self.cores = []

        # 组映射
        self.switch_groups = {}
        self.podid_map = {}  # 设备到podid的映射

    def calculate_hierarchy(self, num_servers):
        """计算三层组网结构"""
        # 计算最大二层支持服务器数量
        # 标准Fat-Tree公式: k^2 / (4 * p)
        # k=交换机端口数, p=每服务器网卡数, 推导: (k/2 * k/2) / p = k^2/(4p)
        max_2tier = (self.switch_ports ** 2) // (4 * self.ports_per_server)

        # 判断是否需要三层组网
        if num_servers <= max_2tier:
            return False, None, None, None

        # 三层组网配置 (Leaf:Spine:Core = 2:2:1)
        # 计算小集群数量 (每个小集群支持 max_2tier 台服务器)
        num_pods = math.ceil(num_servers / max_2tier)

        # 每个小集群的服务器数量
        servers_per_pod = min(max_2tier, num_servers)

        # 每POD的Leaf数量 = groups_per_pod * ports_per_server
        # 每组 servers_per_group 台服务器共享 ports_per_server 个Leaf
        max_servers_per_leaf = self.switch_ports // 2
        servers_per_group = min(servers_per_pod // self.ports_per_server, max_servers_per_leaf)
        groups_per_pod = max(1, servers_per_pod // servers_per_group)
        leafs_per_pod = groups_per_pod * self.ports_per_server

        # Leaf总数 = POD数量 * 每POD Leaf数
        total_leaves = num_pods * leafs_per_pod

        # Spine总数 = Leaf总数 (比例2:2)
        total_spines = total_leaves

        # Core总数 = Spine总数的一半 (比例2:2:1)
        total_cores = max(1, total_spines // 2)

        return True, total_leaves, total_spines, total_cores

    def create_network_objects(self, num_pods, servers_per_pod):
        """创建三层网络对象"""
        # 计算每POD的Leaf数量（与calculate_hierarchy保持一致）
        max_servers_per_leaf = self.switch_ports // 2
        servers_per_group = min(servers_per_pod // self.ports_per_server, max_servers_per_leaf)
        groups_per_pod = max(1, servers_per_pod // servers_per_group)
        leafs_per_pod = groups_per_pod * self.ports_per_server

        # 创建Leaf交换机
        for pod in range(1, num_pods + 1):
            for leaf_idx in range(1, leafs_per_pod + 1):
                leaf_name = f"{self.prefix}Leaf_P{pod}_{leaf_idx}"
                leaf = NetworkObject(
                    name=leaf_name,
                    obj_type=f"{self.network_type}_leaf",
                    group=f"{self.prefix}Leaf组P{pod}",
                    max_ports=self.switch_ports,
                    podid=f"pod-{pod}"  # 使用podid
                )
                self.leaves.append(leaf)
                self.switch_groups[leaf_name] = leaf.group
                self.podid_map[leaf_name] = leaf.podid

        # 创建Spine交换机
        for spine_idx in range(1, len(self.leaves) + 1):
            spine_name = f"{self.prefix}Spine_{spine_idx}"
            spine = NetworkObject(
                name=spine_name,
                obj_type=f"{self.network_type}_spine",
                group=f"{self.prefix}Spine组",
                max_ports=self.switch_ports,
                podid="superpod"  # Spine属于super pod
            )
            self.spines.append(spine)
            self.switch_groups[spine_name] = spine.group
            self.podid_map[spine_name] = spine.podid

        # 创建Core交换机
        core_count = max(1, len(self.spines) // 2)
        for core_idx in range(1, core_count + 1):
            core_name = f"{self.prefix}Core_{core_idx}"
            core = NetworkObject(
                name=core_name,
                obj_type=f"{self.network_type}_core",
                group=f"{self.prefix}Core组",
                max_ports=self.switch_ports,
                podid="superpod"  # Core属于super pod
            )
            self.cores.append(core)
            self.switch_groups[core_name] = core.group
            self.podid_map[core_name] = core.podid

    def generate_connections(self, servers, num_pods, servers_per_pod):
        """生成三层网络连接关系，使用动态端口分配策略"""
        connections = []

        # 1. 服务器到Leaf连接 - 使用Leaf的下联端口
        # 计算每个Leaf可连接的服务器数量（前50%端口）
        max_servers_per_leaf = math.floor(self.switch_ports / 2)
        servers_per_group = min(servers_per_pod // self.ports_per_server, max_servers_per_leaf)

        if servers_per_group <= 0:
            servers_per_group = 1

        for server in servers:
            server_idx = int(server.name.split('_')[1])
            pod_id = (server_idx - 1) // servers_per_pod + 1
            server.podid = f"pod-{pod_id}"

            # 服务器在POD内的索引 (0-based)
            server_idx_in_pod = (server_idx - 1) % servers_per_pod

            # 确定服务器所属的组
            group_index = server_idx_in_pod // servers_per_group

            for port_idx in range(1, self.ports_per_server + 1):
                # 计算Leaf索引 (1-based)
                leaf_idx = group_index * self.ports_per_server + port_idx
                leaf_name = f"{self.prefix}Leaf_P{pod_id}_{leaf_idx}"

                leaf = next((l for l in self.leaves if l.name == leaf_name), None)
                if not leaf:
                    continue

                # 检查Leaf下联端口是否可用
                if leaf.downlink_counter > leaf.downlink_limit:
                    # 查找同POD内其他可用Leaf
                    backup_found = False
                    for backup_idx in range(1, len(self.leaves) // num_pods + 1):
                        if backup_idx == leaf_idx:
                            continue
                        backup_name = f"{self.prefix}Leaf_P{pod_id}_{backup_idx}"
                        backup_leaf = next((l for l in self.leaves if l.name == backup_name), None)
                        if backup_leaf and backup_leaf.downlink_counter <= backup_leaf.downlink_limit:
                            leaf = backup_leaf
                            backup_found = True
                            break

                    if not backup_found:
                        raise ValueError(f"{leaf_name}下联端口不足，且无可用备用Leaf")

                server_port = f"{self.prefix}网卡{port_idx}"
                try:
                    leaf_port = leaf.get_downlink_port()
                except ValueError as e:
                    print(f"警告: {str(e)}")
                    continue

                # 服务器到Leaf
                conn_down = Connection(
                    a_device=server.name,
                    a_port=server_port,
                    a_module=self.network_speed,
                    z_device=leaf.name,
                    z_port=leaf_port,
                    z_module=self.network_speed,
                    cable_type=self.cable_type_config['server_leaf'],
                    description=f"服务器到{self.prefix}Leaf"
                )

                # Leaf到服务器
                conn_up = Connection(
                    a_device=leaf.name,
                    a_port=leaf_port,
                    a_module=self.network_speed,
                    z_device=server.name,
                    z_port=server_port,
                    z_module=self.network_speed,
                    cable_type=self.cable_type_config['server_leaf'],
                    description=f"{self.prefix}Leaf到服务器"
                )

                server.add_connection(conn_down)
                leaf.add_connection(conn_up)
                connections.extend([conn_down, conn_up])

        # 2. Leaf到Spine连接 - Leaf使用上联端口，Spine使用下联端口
        # 使用轮转(Round-Robin)方式分配Leaf→Spine连接，防止Spine端口溢出
        for leaf in self.leaves:
            # 获取Leaf所在的POD
            pod_id = int(re.search(r"P(\d+)", leaf.name).group(1))

            # 计算该POD分配的Spine范围
            start_spine = (pod_id - 1) * len(self.spines) // num_pods + 1
            end_spine = pod_id * len(self.spines) // num_pods
            spines_for_pod = end_spine - start_spine + 1

            # 计算每个Leaf的最大上联数
            max_uplinks = leaf.uplink_limit - leaf.uplink_counter + 1
            connections_per_leaf = min(max_uplinks, spines_for_pod)

            # 轮转偏移：POD内第i个Leaf从第 (i * connections_per_leaf) % spines_for_pod 个Spine开始连接
            leaf_idx_in_pod = int(re.search(r'_(\d+)$', leaf.name).group(1)) - 1
            spine_offset = (leaf_idx_in_pod * connections_per_leaf) % spines_for_pod

            for i in range(connections_per_leaf):
                spine_idx = start_spine + (spine_offset + i) % spines_for_pod
                spine_name = f"{self.prefix}Spine_{spine_idx}"
                spine = next((s for s in self.spines if s.name == spine_name), None)
                if not spine:
                    continue

                try:
                    leaf_port = leaf.get_uplink_port()
                    spine_port = spine.get_downlink_port()

                    # Leaf到Spine
                    conn_leaf_to_spine = Connection(
                        a_device=leaf.name,
                        a_port=leaf_port,
                        a_module=self.network_speed,
                        z_device=spine.name,
                        z_port=spine_port,
                        z_module=self.network_speed,
                        cable_type=self.cable_type_config['leaf_spine'],
                        description=f"{self.prefix}Leaf到Spine"
                    )

                    # Spine到Leaf
                    conn_spine_to_leaf = Connection(
                        a_device=spine.name,
                        a_port=spine_port,
                        a_module=self.network_speed,
                        z_device=leaf.name,
                        z_port=leaf_port,
                        z_module=self.network_speed,
                        cable_type=self.cable_type_config['leaf_spine'],
                        description=f"{self.prefix}Spine到Leaf"
                    )

                    leaf.add_connection(conn_leaf_to_spine)
                    spine.add_connection(conn_spine_to_leaf)
                    connections.extend([conn_leaf_to_spine, conn_spine_to_leaf])

                except ValueError as e:
                    print(f"警告: {str(e)}")
                    continue

        # 3. Spine到Core连接 - Spine使用上联端口，Core使用所有端口
        # 使用轮转(Round-Robin)方式分配Spine→Core连接，防止Core端口溢出
        if self.cores:
            # 每个Spine的最大上联端口数
            max_uplinks_per_spine = self.switch_ports // 2

            # 计算每个Core可接受的连接数
            max_links_per_core = [core.core_limit for core in self.cores]
            # 使用计数器跟踪每个Core已用端口
            core_port_used = [0] * len(self.cores)

            # 为每个Spine分配Core连接
            for spine_idx, spine in enumerate(self.spines):
                # 计算Spine剩余上联端口
                remaining_uplinks = spine.uplink_limit - spine.uplink_counter + 1
                if remaining_uplinks <= 0:
                    continue

                # 确定要分配的上行连接数
                uplinks = min(remaining_uplinks, max_uplinks_per_spine)

                # 轮转偏移：第s个Spine从第 (s * uplinks) % len(self.cores) 个Core开始连接
                core_offset = (spine_idx * uplinks) % len(self.cores)

                for i in range(uplinks):
                    core_idx = (core_offset + i) % len(self.cores)
                    if core_port_used[core_idx] >= max_links_per_core[core_idx]:
                        continue

                    try:
                        spine_port = spine.get_uplink_port()
                        core_port = self.cores[core_idx].get_core_port()

                        # 更新Core可用端口计数
                        core_port_used[core_idx] += 1

                        # Spine到Core
                        conn_spine_to_core = Connection(
                            a_device=spine.name,
                            a_port=spine_port,
                            a_module=self.network_speed,
                            z_device=self.cores[core_idx].name,
                            z_port=core_port,
                            z_module=self.network_speed,
                            cable_type=self.cable_type_config['spine_core'],
                            description=f"{self.prefix}Spine到Core"
                        )

                        # Core到Spine
                        conn_core_to_spine = Connection(
                            a_device=self.cores[core_idx].name,
                            a_port=core_port,
                            a_module=self.network_speed,
                            z_device=spine.name,
                            z_port=spine_port,
                            z_module=self.network_speed,
                            cable_type=self.cable_type_config['spine_core'],
                            description=f"{self.prefix}Core到Spine"
                        )

                        spine.add_connection(conn_spine_to_core)
                        self.cores[core_idx].add_connection(conn_core_to_spine)
                        connections.extend([conn_spine_to_core, conn_core_to_spine])

                    except ValueError as e:
                        print(f"警告: {str(e)}")
                        continue

        return connections


class NetworkDesigner:
    def __init__(self, config_file="network_config.ini"):
        # 加载配置文件
        self.config = configparser.ConfigParser()
        if os.path.exists(config_file):
            # 明确指定 UTF-8 编码
            with open(config_file, 'r', encoding='utf-8') as f:
                self.config.read_file(f)
        else:
            # 默认配置...
            self.config['DEFAULT'] = {
                'num_servers': '256',
                'param_ports_per_server': '8',
                'storage_ports_per_server': '1',
                'param_switch_ports': '64',
                'storage_switch_ports': '40',
                'param_speed': '400G',
                'storage_speed': '200G',
                'cable_param_server_leaf': 'MPO',
                'cable_param_leaf_spine': 'MPO',
                'cable_param_spine_core': 'MPO',
                'cable_storage_server_leaf': 'MPO',
                'cable_storage_leaf_spine': 'MPO',
                'cable_storage_spine_core': 'MPO'
            }

        # 解析配置
        self.num_servers = int(self.config.get('DEFAULT', 'num_servers', fallback=256))
        self.param_ports_per_server = int(self.config.get('DEFAULT', 'param_ports_per_server', fallback=8))
        self.storage_ports_per_server = int(self.config.get('DEFAULT', 'storage_ports_per_server', fallback=1))
        self.param_switch_ports = int(self.config.get('DEFAULT', 'param_switch_ports', fallback=64))
        self.storage_switch_ports = int(self.config.get('DEFAULT', 'storage_switch_ports', fallback=40))
        self.param_speed = self.config.get('DEFAULT', 'param_speed', fallback="400G")
        self.storage_speed = self.config.get('DEFAULT', 'storage_speed', fallback="200G")

        # 线缆类型配置
        self.cable_types = {
            'param': {
                'server_leaf': self.config.get('DEFAULT', 'cable_param_server_leaf', fallback='MPO'),
                'leaf_spine': self.config.get('DEFAULT', 'cable_param_leaf_spine', fallback='MPO'),
                'spine_core': self.config.get('DEFAULT', 'cable_param_spine_core', fallback='MPO')
            },
            'storage': {
                'server_leaf': self.config.get('DEFAULT', 'cable_storage_server_leaf', fallback='MPO'),
                'leaf_spine': self.config.get('DEFAULT', 'cable_storage_leaf_spine', fallback='MPO'),
                'spine_core': self.config.get('DEFAULT', 'cable_storage_spine_core', fallback='MPO')
            }
        }

        # 网络对象存储
        self.servers = []
        self.param_leaves = []
        self.param_spines = []
        self.param_cores = []
        self.storage_leaves = []
        self.storage_spines = []
        self.storage_cores = []

        # 组映射
        self.server_groups = {}
        self.switch_groups = {}
        self.podid_map = {}  # 设备到podid的映射

        # 计算网络层次
        self.calc_network_hierarchy()

        # 创建网络对象
        self.create_network_objects()

        # 生成连接
        self.generate_connections()

    def calc_network_hierarchy(self):
        """计算网络层次结构（参数和存储网络）"""
        # 参数网络
        param_3tier = ThreeTierNetwork(
            self.param_ports_per_server,
            self.param_switch_ports,
            self.param_speed,
            self.cable_types['param'],
            "param"
        )

        self.param_3tier_needed, param_leaves, param_spines, param_cores = param_3tier.calculate_hierarchy(
            self.num_servers)

        if self.param_3tier_needed:
            # 三层组网
            self.param_leaf_count = param_leaves
            self.param_spine_count = param_spines
            self.param_core_count = param_cores

            # 计算POD数量 (使用标准Fat-Tree公式: k^2/(4p))
            max_2tier = (self.param_switch_ports ** 2) // (4 * self.param_ports_per_server)
            self.param_pods = math.ceil(self.num_servers / max_2tier)
            self.param_servers_per_pod = min(max_2tier, self.num_servers)
        else:
            # 二层组网
            self.param_servers_per_group = min(self.param_switch_ports // 2, self.num_servers)
            self.param_groups = math.ceil(self.num_servers / self.param_servers_per_group)
            self.param_leaf_per_group = self.param_ports_per_server
            self.param_leaf_count = self.param_groups * self.param_leaf_per_group
            self.param_spine_count = max(1, self.param_leaf_count // 2)
            self.param_core_count = 0
            self.param_pods = 0
            self.param_servers_per_pod = 0  # 初始化避免后续错误

        # 存储网络
        storage_3tier = ThreeTierNetwork(
            self.storage_ports_per_server,
            self.storage_switch_ports,
            self.storage_speed,
            self.cable_types['storage'],
            "storage"
        )

        self.storage_3tier_needed, storage_leaves, storage_spines, storage_cores = storage_3tier.calculate_hierarchy(
            self.num_servers)

        if self.storage_3tier_needed:
            # 三层组网
            self.storage_leaf_count = storage_leaves
            self.storage_spine_count = storage_spines
            self.storage_core_count = storage_cores

            # 计算POD数量 (使用标准Fat-Tree公式: k^2/(4p))
            max_2tier = (self.storage_switch_ports ** 2) // (4 * self.storage_ports_per_server)
            self.storage_pods = math.ceil(self.num_servers / max_2tier)
            self.storage_servers_per_pod = min(max_2tier, self.num_servers)
        else:
            # 二层组网 - 独立计算，不依赖参数网络的POD结构
            self.storage_servers_per_group = min(self.storage_switch_ports // 2, self.num_servers)
            self.storage_groups = math.ceil(self.num_servers / self.storage_servers_per_group)

            self.storage_leaf_per_group = self.storage_ports_per_server
            self.storage_leaf_count = self.storage_groups * self.storage_leaf_per_group
            self.storage_spine_count = max(1, self.storage_leaf_count // 2)
            self.storage_core_count = 0
            self.storage_pods = 0
            self.storage_servers_per_pod = 0  # 初始化避免后续错误

    def create_network_objects(self):
        """创建网络对象"""
        # 创建服务器
        for server_idx in range(1, self.num_servers + 1):
            # 确定组/POD
            if self.param_3tier_needed:
                pod_id = (server_idx - 1) // self.param_servers_per_pod + 1
                group_name = f"服务器POD{pod_id}"
                podid = f"pod-{pod_id}"
            else:
                group_id = (server_idx - 1) // self.param_servers_per_group + 1
                group_name = f"服务器组{group_id}"
                podid = f"pod-{group_id}"

            server = NetworkObject(
                name=f"GPU服务器_{server_idx}",
                obj_type='server',
                group=group_name,
                podid=podid  # 设置服务器podid
            )
            self.servers.append(server)
            self.server_groups[server.name] = group_name
            self.podid_map[server.name] = podid

        # 参数网络交换机
        if self.param_3tier_needed:
            # 三层组网
            param_3tier = ThreeTierNetwork(
                self.param_ports_per_server,
                self.param_switch_ports,
                self.param_speed,
                self.cable_types['param'],
                "param"
            )
            param_3tier.create_network_objects(self.param_pods, self.param_servers_per_pod)
            self.param_leaves = param_3tier.leaves
            self.param_spines = param_3tier.spines
            self.param_cores = param_3tier.cores
            self.switch_groups.update(param_3tier.switch_groups)
            self.podid_map.update(param_3tier.podid_map)
        else:
            # 二层组网
            for group in range(1, self.param_groups + 1):
                for leaf_idx in range(1, self.param_leaf_per_group + 1):
                    leaf_name = f"参数Leaf_G{group}_{leaf_idx}"
                    leaf = NetworkObject(
                        name=leaf_name,
                        obj_type='param_leaf',
                        group=f"参数Leaf组{group}",
                        max_ports=self.param_switch_ports,
                        podid=f"pod-{group}"  # 二层组网的podid
                    )
                    self.param_leaves.append(leaf)
                    self.switch_groups[leaf_name] = leaf.group
                    self.podid_map[leaf_name] = leaf.podid

            for spine_idx in range(1, self.param_spine_count + 1):
                spine_name = f"参数Spine_{spine_idx}"
                spine = NetworkObject(
                    name=spine_name,
                    obj_type='param_spine',
                    group="参数Spine组",
                    max_ports=self.param_switch_ports,
                    podid="superpod"  # Spine属于super pod
                )
                self.param_spines.append(spine)
                self.switch_groups[spine_name] = spine.group
                self.podid_map[spine_name] = spine.podid

        # 存储网络交换机
        if self.storage_3tier_needed:
            # 三层组网
            storage_3tier = ThreeTierNetwork(
                self.storage_ports_per_server,
                self.storage_switch_ports,
                self.storage_speed,
                self.cable_types['storage'],
                "storage"
            )
            storage_3tier.create_network_objects(self.storage_pods, self.storage_servers_per_pod)
            self.storage_leaves = storage_3tier.leaves
            self.storage_spines = storage_3tier.spines
            self.storage_cores = storage_3tier.cores
            self.switch_groups.update(storage_3tier.switch_groups)
            self.podid_map.update(storage_3tier.podid_map)
        else:
            # 二层组网
            for group in range(1, self.storage_groups + 1):
                for leaf_idx in range(1, self.storage_leaf_per_group + 1):
                    leaf_name = f"存储Leaf_G{group}_{leaf_idx}"
                    leaf = NetworkObject(
                        name=leaf_name,
                        obj_type='storage_leaf',
                        group=f"存储Leaf组{group}",
                        max_ports=self.storage_switch_ports,
                        podid=f"pod-{group}"  # 二层组网的podid
                    )
                    self.storage_leaves.append(leaf)
                    self.switch_groups[leaf_name] = leaf.group
                    self.podid_map[leaf_name] = leaf.podid

            for spine_idx in range(1, self.storage_spine_count + 1):
                spine_name = f"存储Spine_{spine_idx}"
                spine = NetworkObject(
                    name=spine_name,
                    obj_type='storage_spine',
                    group="存储Spine组",
                    max_ports=self.storage_switch_ports,
                    podid="superpod"  # Spine属于super pod
                )
                self.storage_spines.append(spine)
                self.switch_groups[spine_name] = spine.group
                self.podid_map[spine_name] = spine.podid

    def generate_connections(self):
        """生成网络连接"""
        # 参数网络连接
        if self.param_3tier_needed:
            param_3tier = ThreeTierNetwork(
                self.param_ports_per_server,
                self.param_switch_ports,
                self.param_speed,
                self.cable_types['param'],
                "param"
            )
            param_3tier.leaves = self.param_leaves
            param_3tier.spines = self.param_spines
            param_3tier.cores = self.param_cores
            param_3tier.generate_connections(self.servers, self.param_pods, self.param_servers_per_pod)
        else:
            # 二层组网连接逻辑
            self._generate_param_2tier_connections()

        # 存储网络连接
        if self.storage_3tier_needed:
            storage_3tier = ThreeTierNetwork(
                self.storage_ports_per_server,
                self.storage_switch_ports,
                self.storage_speed,
                self.cable_types['storage'],
                "storage"
            )
            storage_3tier.leaves = self.storage_leaves
            storage_3tier.spines = self.storage_spines
            storage_3tier.cores = self.storage_cores
            storage_3tier.generate_connections(self.servers, self.storage_pods, self.storage_servers_per_pod)
        else:
            # 二层组网连接逻辑
            self._generate_storage_2tier_connections()

    def _generate_param_2tier_connections(self):
        """生成参数网络二层组网连接 - 使用动态端口分配策略"""
        # 1. 服务器到参数Leaf连接 - 使用Leaf的下联端口
        for server in self.servers:
            server_idx = int(server.name.split('_')[1])
            group_id = (server_idx - 1) // self.param_servers_per_group + 1

            for port_idx in range(1, self.param_ports_per_server + 1):
                leaf_idx = port_idx
                leaf_name = f"参数Leaf_G{group_id}_{leaf_idx}"
                leaf = next((l for l in self.param_leaves if l.name == leaf_name), None)
                if not leaf:
                    continue

                try:
                    server_port = f"参数网卡{port_idx}"
                    leaf_port = leaf.get_downlink_port()  # Leaf下联端口

                    # 服务器到Leaf
                    conn_down = Connection(
                        a_device=server.name,
                        a_port=server_port,
                        a_module=self.param_speed,
                        z_device=leaf.name,
                        z_port=leaf_port,
                        z_module=self.param_speed,
                        cable_type=self.cable_types['param']['server_leaf'],
                        description="服务器到参数Leaf"
                    )

                    # Leaf到服务器
                    conn_up = Connection(
                        a_device=leaf.name,
                        a_port=leaf_port,
                        a_module=self.param_speed,
                        z_device=server.name,
                        z_port=server_port,
                        z_module=self.param_speed,
                        cable_type=self.cable_types['param']['server_leaf'],
                        description="参数Leaf到服务器"
                    )

                    server.add_connection(conn_down)
                    leaf.add_connection(conn_up)

                except ValueError as e:
                    print(f"警告: {str(e)}")
                    continue

        # 2. 参数Leaf到Spine连接 - Leaf使用上联端口，Spine使用下联端口
        for leaf in self.param_leaves:
            # 每个Leaf需要连接的Spine数量（不超过上联端口数量）
            max_uplinks = leaf.uplink_limit - leaf.uplink_counter + 1
            spines_per_leaf = min(max_uplinks, self.param_spine_count)

            for spine_idx in range(1, spines_per_leaf + 1):
                spine_name = f"参数Spine_{spine_idx}"
                spine = next((s for s in self.param_spines if s.name == spine_name), None)
                if not spine:
                    continue

                try:
                    leaf_port = leaf.get_uplink_port()  # Leaf上联端口
                    spine_port = spine.get_downlink_port()  # Spine下联端口

                    # Leaf到Spine
                    conn_leaf_to_spine = Connection(
                        a_device=leaf.name,
                        a_port=leaf_port,
                        a_module=self.param_speed,
                        z_device=spine.name,
                        z_port=spine_port,
                        z_module=self.param_speed,
                        cable_type=self.cable_types['param']['leaf_spine'],
                        description="参数Leaf到Spine"
                    )

                    # Spine到Leaf
                    conn_spine_to_leaf = Connection(
                        a_device=spine.name,
                        a_port=spine_port,
                        a_module=self.param_speed,
                        z_device=leaf.name,
                        z_port=leaf_port,
                        z_module=self.param_speed,
                        cable_type=self.cable_types['param']['leaf_spine'],
                        description="参数Spine到Leaf"
                    )

                    leaf.add_connection(conn_leaf_to_spine)
                    spine.add_connection(conn_spine_to_leaf)

                except ValueError as e:
                    print(f"警告: {str(e)}")
                    continue

    def _generate_storage_2tier_connections(self):
        """生成存储网络二层组网连接 - 使用动态端口分配策略"""
        # 1. 服务器到存储Leaf连接 - 使用Leaf的下联端口
        for server in self.servers:
            server_idx = int(server.name.split('_')[1])
            group_id = (server_idx - 1) // self.storage_servers_per_group + 1

            for port_idx in range(1, self.storage_ports_per_server + 1):
                leaf_idx = port_idx
                leaf_name = f"存储Leaf_G{group_id}_{leaf_idx}"
                leaf = next((l for l in self.storage_leaves if l.name == leaf_name), None)
                if not leaf:
                    continue

                try:
                    server_port = f"存储网卡{port_idx}"
                    leaf_port = leaf.get_downlink_port()  # Leaf下联端口

                    # 服务器到Leaf
                    conn_down = Connection(
                        a_device=server.name,
                        a_port=server_port,
                        a_module=self.storage_speed,
                        z_device=leaf.name,
                        z_port=leaf_port,
                        z_module=self.storage_speed,
                        cable_type=self.cable_types['storage']['server_leaf'],
                        description="服务器到存储"
                    )

                    # Leaf到服务器
                    conn_up = Connection(
                        a_device=leaf.name,
                        a_port=leaf_port,
                        a_module=self.storage_speed,
                        z_device=server.name,
                        z_port=server_port,
                        z_module=self.storage_speed,
                        cable_type=self.cable_types['storage']['server_leaf'],
                        description="存储下行连接"
                    )

                    server.add_connection(conn_down)
                    leaf.add_connection(conn_up)

                except ValueError as e:
                    print(f"警告: {str(e)}")
                    continue

        # 2. 存储Leaf到Spine连接 - Leaf使用上联端口，Spine使用下联端口
        for leaf in self.storage_leaves:
            # 每个Leaf需要连接的Spine数量（不超过上联端口数量）
            max_uplinks = leaf.uplink_limit - leaf.uplink_counter + 1
            spines_per_leaf = min(max_uplinks, self.storage_spine_count)

            for spine_idx in range(1, spines_per_leaf + 1):
                spine_name = f"存储Spine_{spine_idx}"
                spine = next((s for s in self.storage_spines if s.name == spine_name), None)
                if not spine:
                    continue

                try:
                    leaf_port = leaf.get_uplink_port()  # Leaf上联端口
                    spine_port = spine.get_downlink_port()  # Spine下联端口

                    # Leaf到Spine
                    conn_leaf_to_spine = Connection(
                        a_device=leaf.name,
                        a_port=leaf_port,
                        a_module=self.storage_speed,
                        z_device=spine.name,
                        z_port=spine_port,
                        z_module=self.storage_speed,
                        cable_type=self.cable_types['storage']['leaf_spine'],
                        description="存储网络上联"
                    )

                    # Spine到Leaf
                    conn_spine_to_leaf = Connection(
                        a_device=spine.name,
                        a_port=spine_port,
                        a_module=self.storage_speed,
                        z_device=leaf.name,
                        z_port=leaf_port,
                        z_module=self.storage_speed,
                        cable_type=self.cable_types['storage']['leaf_spine'],
                        description="存储网络下联"
                    )

                    leaf.add_connection(conn_leaf_to_spine)
                    spine.add_connection(conn_spine_to_leaf)

                except ValueError as e:
                    print(f"警告: {str(e)}")
                    continue

    def extract_number(self, s):
        """从字符串中提取数字用于排序"""
        if isinstance(s, (int, float)):
            return int(s)
        if not isinstance(s, str):
            return 0
        match = re.search(r'\d+', s)
        return int(match.group()) if match else 0

    def get_switch_type_weight(self, device_name):
        """获取交换机类型权重用于排序"""
        if "Leaf" in device_name:
            return 1
        elif "Spine" in device_name:
            return 2
        elif "Core" in device_name:
            return 3
        return 4

    def generate_server_view(self):
        """生成服务器视角的连接表"""
        connections = []

        for server in self.servers:
            server_group = self.server_groups[server.name]
            podid = self.podid_map.get(server.name, "")
            for conn in server.connections:
                if conn.a_device == server.name:
                    connections.append({
                        'podid': podid,  # 第一列：podid
                        '服务器分组': server_group,
                        'A端设备': conn.a_device,
                        'A端接口': conn.a_port,
                        'A端模块': conn.a_module,
                        'Z端设备': conn.z_device,
                        'Z端接口': conn.z_port,
                        'Z端模块': conn.z_module,
                        '线缆类型': conn.cable_type,
                        '描述': conn.description
                    })

        df = pd.DataFrame(connections)
        if not df.empty:
            df['group_num'] = df['服务器分组'].apply(self.extract_number)
            df['server_num'] = df['A端设备'].apply(self.extract_number)
            df = df.sort_values(by=['group_num', 'server_num', 'A端接口'])
            df = df.drop(columns=['group_num', 'server_num'])
        return df

    def generate_switch_view(self):
        """生成交换机视角的连接表"""
        param_connections = []
        storage_connections = []

        all_switches = self.param_leaves + self.param_spines + self.param_cores + \
                       self.storage_leaves + self.storage_spines + self.storage_cores

        for switch in all_switches:
            switch_group = self.switch_groups.get(switch.name, "")
            podid = self.podid_map.get(switch.name, "")
            for conn in switch.connections:
                if conn.a_device == switch.name:
                    row = {
                        'podid': podid,  # 第一列：podid
                        '交换机分组': switch_group,
                        'A端设备': conn.a_device,
                        'A端接口': conn.a_port,
                        'A端模块': conn.a_module,
                        'Z端设备': conn.z_device,
                        'Z端接口': conn.z_port,
                        'Z端模块': conn.z_module,
                        '线缆类型': conn.cable_type,
                        '描述': conn.description
                    }
                    if "参数" in switch.name:
                        param_connections.append(row)
                    else:
                        storage_connections.append(row)

        param_df = pd.DataFrame(param_connections)
        storage_df = pd.DataFrame(storage_connections)

        # 排序参数网络连接
        if not param_df.empty:
            param_df['type_weight'] = param_df['A端设备'].apply(self.get_switch_type_weight)
            param_df['group_num'] = param_df['A端设备'].apply(
                lambda x: int(re.search(r'[PG](\d+)', x).group(1)) if re.search(r'[PG](\d+)', x) else 0
            )
            param_df['device_num'] = param_df['A端设备'].apply(
                lambda x: int(re.search(r'_(\d+)$', x).group(1)) if re.search(r'_(\d+)$', x) else 0
            )
            param_df['port_num'] = param_df['A端接口'].apply(self.extract_number)
            param_df = param_df.sort_values(by=['type_weight', 'group_num', 'device_num', 'port_num'])
            param_df = param_df.drop(columns=['type_weight', 'group_num', 'device_num', 'port_num'])

        # 排序存储网络连接
        if not storage_df.empty:
            storage_df['type_weight'] = storage_df['A端设备'].apply(self.get_switch_type_weight)
            storage_df['group_num'] = storage_df['A端设备'].apply(
                lambda x: int(re.search(r'[PG](\d+)', x).group(1)) if re.search(r'[PG](\d+)', x) else 0
            )
            storage_df['device_num'] = storage_df['A端设备'].apply(
                lambda x: int(re.search(r'_(\d+)$', x).group(1)) if re.search(r'_(\d+)$', x) else 0
            )
            storage_df['port_num'] = storage_df['A端接口'].apply(self.extract_number)
            storage_df = storage_df.sort_values(by=['type_weight', 'group_num', 'device_num', 'port_num'])
            storage_df = storage_df.drop(columns=['type_weight', 'group_num', 'device_num', 'port_num'])

        return {
            '参数网络': param_df,
            '存储网络': storage_df
        }

    def generate_summary_data(self):
        """生成网络设计摘要数据"""
        # 计算参数网络端口使用率（下联/上联各占switch_ports//2）
        half_ports = self.param_switch_ports // 2
        if self.param_3tier_needed:
            servers_per_group = min(self.param_servers_per_pod // self.param_ports_per_server, half_ports)
            param_downlink_usage = f"{servers_per_group}/{half_ports} ({servers_per_group / half_ports * 100:.1f}%)"
            param_uplink_usage = f"{half_ports}/{half_ports} (100.0%)"
        else:
            param_downlink_usage = f"{self.param_servers_per_group}/{half_ports} ({self.param_servers_per_group / half_ports * 100:.1f}%)"
            uplinks_used = min(half_ports, self.param_spine_count)
            param_uplink_usage = f"{uplinks_used}/{half_ports} ({uplinks_used / half_ports * 100:.1f}%)"

        # 计算存储网络端口使用率
        half_ports = self.storage_switch_ports // 2
        if self.storage_3tier_needed:
            servers_per_group = min(self.storage_servers_per_pod // self.storage_ports_per_server, half_ports)
            storage_downlink_usage = f"{servers_per_group}/{half_ports} ({servers_per_group / half_ports * 100:.1f}%)"
            storage_uplink_usage = f"{half_ports}/{half_ports} (100.0%)"
        else:
            storage_downlink_usage = f"{self.storage_servers_per_group}/{half_ports} ({self.storage_servers_per_group / half_ports * 100:.1f}%)"
            uplinks_used = min(half_ports, self.storage_spine_count)
            storage_uplink_usage = f"{uplinks_used}/{half_ports} ({uplinks_used / half_ports * 100:.1f}%)"

        # 三层组网信息
        param_tier_info = "3层(Leaf-Spine-Core)" if self.param_3tier_needed else "2层(Leaf-Spine)"
        storage_tier_info = "3层(Leaf-Spine-Core)" if self.storage_3tier_needed else "2层(Leaf-Spine)"

        # 分组信息
        if self.param_3tier_needed:
            param_group_info = f"{self.param_pods}个POD, 每个POD{self.param_servers_per_pod}台"
        else:
            param_group_info = f"{self.param_groups}组, 每组{self.param_servers_per_group}台"

        if self.storage_3tier_needed:
            storage_group_info = f"{self.storage_pods}个POD, 每个POD{self.storage_servers_per_pod}台"
        else:
            storage_group_info = f"{self.storage_groups}组, 每组{self.storage_servers_per_group}台"

        summary = [
            ["网络设计摘要", ""],
            ["设计时间", pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["", ""],
            ["GPU服务器配置", ""],
            ["服务器数量", self.num_servers],
            ["参数网卡/服务器", self.param_ports_per_server],
            ["存储网卡/服务器", self.storage_ports_per_server],
            ["", ""],
            ["参数网络设计", ""],
            ["交换机端口数", self.param_switch_ports],
            ["Leaf交换机数量", self.param_leaf_count],
            ["Spine交换机数量", self.param_spine_count],
            ["Core交换机数量", self.param_core_count if self.param_3tier_needed else "无"],
            ["网络层级", param_tier_info],
            ["服务器分组", param_group_info],
            ["下行端口使用率", param_downlink_usage],
            ["上行端口使用率", param_uplink_usage],
            ["收敛比例", "1:1:1" if self.param_3tier_needed else "1:1"],
            ["", ""],
            ["存储网络设计", ""],
            ["交换机端口数", self.storage_switch_ports],
            ["Leaf交换机数量", self.storage_leaf_count],
            ["Spine交换机数量", self.storage_spine_count],
            ["Core交换机数量", self.storage_core_count if self.storage_3tier_needed else "无"],
            ["网络层级", storage_tier_info],
            ["服务器分组", storage_group_info],
            ["下行端口使用率", storage_downlink_usage],
            ["上行端口使用率", storage_uplink_usage],
            ["收敛比例", "1:1:1" if self.storage_3tier_needed else "1:1"],
            ["", ""],
            ["网络速度配置", ""],
            ["参数网络速度", self.param_speed],
            ["存储网络速度", self.storage_speed],
            ["线缆类型", "MPO（多模光纤）"]
        ]
        return pd.DataFrame(summary, columns=["项目", "值"])

    def apply_excel_formatting(self, filename):
        """应用Excel格式美化：调整列宽、合并单元格、居中、边框等"""
        wb = load_workbook(filename)
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_font = Font(bold=True, size=12, color="FFFFFF")
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

        # 列宽配置
        col_widths = {
            'podid': 15,  # podid列宽度
            '服务器分组': 15,
            '交换机分组': 15,
            'A端设备': 30,
            'A端接口': 15,
            'A端模块': 12,
            'Z端设备': 30,
            'Z端接口': 15,
            'Z端模块': 12,
            '线缆类型': 12,
            '描述': 30,
            '项目': 25,
            '值': 40
        }

        # 处理每个sheet
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # 设置冻结首行
            if ws.max_row > 1:
                ws.freeze_panes = 'A2'

            # 设置列宽
            for col_idx, col_name in enumerate(ws[1], 1):
                col_letter = get_column_letter(col_idx)
                col_name_value = col_name.value
                if col_name_value in col_widths:
                    ws.column_dimensions[col_letter].width = col_widths[col_name_value]
                else:
                    ws.column_dimensions[col_letter].width = 15

            # 设置行高
            ws.row_dimensions[1].height = 25  # 标题行高度

            # 设置标题行样式
            for row in ws.iter_rows(min_row=1, max_row=1):
                for cell in row:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_align
                    cell.border = thin_border

            # 确定podid列位置
            podid_col = None
            for col_idx, cell in enumerate(ws[1], 1):
                if cell.value == 'podid':
                    podid_col = col_idx
                    break

            # 确定分组列位置
            group_col = None
            for col_idx, cell in enumerate(ws[1], 1):
                if cell.value in ['服务器分组', '交换机分组']:
                    group_col = col_idx
                    break

            # 确定设备列位置
            device_col = None
            for col_idx, cell in enumerate(ws[1], 1):
                if cell.value == 'A端设备':
                    device_col = col_idx
                    break

            # 记录合并区域
            merge_regions = []

            # 处理网络设计摘要的特殊格式
            if sheet_name == '网络设计摘要':
                for row_idx in range(1, ws.max_row + 1):
                    for col_idx in range(1, ws.max_column + 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.alignment = center_align
                        cell.border = thin_border
                continue

            # podid列合并
            if podid_col:
                current_podid = None
                start_row = 2

                for row_idx in range(2, ws.max_row + 1):
                    podid_cell = ws.cell(row=row_idx, column=podid_col)
                    podid_value = podid_cell.value

                    if podid_value != current_podid:
                        if current_podid is not None:
                            if row_idx - 1 > start_row:
                                merge_regions.append((start_row, row_idx - 1, podid_col, podid_col))
                        current_podid = podid_value
                        start_row = row_idx

                # 合并最后一个podid
                if current_podid is not None and ws.max_row >= start_row:
                    if ws.max_row > start_row:
                        merge_regions.append((start_row, ws.max_row, podid_col, podid_col))

            # 分组列合并
            if group_col:
                current_group = None
                start_row = 2

                for row_idx in range(2, ws.max_row + 1):
                    group_cell = ws.cell(row=row_idx, column=group_col)
                    group_value = group_cell.value

                    if group_value != current_group:
                        if current_group is not None:
                            if row_idx - 1 > start_row:
                                merge_regions.append((start_row, row_idx - 1, group_col, group_col))
                        current_group = group_value
                        start_row = row_idx

                # 合并最后一个分组
                if current_group is not None and ws.max_row >= start_row:
                    if ws.max_row > start_row:
                        merge_regions.append((start_row, ws.max_row, group_col, group_col))

            # 设备列合并
            if device_col:
                current_device = None
                start_row = 2

                for row_idx in range(2, ws.max_row + 1):
                    device_cell = ws.cell(row=row_idx, column=device_col)
                    device_value = device_cell.value

                    if group_col:
                        # 检查分组是否变化
                        group_cell = ws.cell(row=row_idx, column=group_col)
                        prev_group_cell = ws.cell(row=row_idx - 1, column=group_col) if row_idx > 2 else None

                        if prev_group_cell and group_cell.value != prev_group_cell.value:
                            if current_device is not None:
                                if row_idx - 1 > start_row:
                                    merge_regions.append((start_row, row_idx - 1, device_col, device_col))
                            current_device = None

                    if device_value != current_device:
                        if current_device is not None:
                            if row_idx - 1 > start_row:
                                merge_regions.append((start_row, row_idx - 1, device_col, device_col))
                        current_device = device_value
                        start_row = row_idx

                # 合并最后一个设备
                if current_device is not None and ws.max_row >= start_row:
                    if ws.max_row > start_row:
                        merge_regions.append((start_row, ws.max_row, device_col, device_col))

            # 应用合并
            for region in merge_regions:
                start_row, end_row, start_col, end_col = region
                ws.merge_cells(start_row=start_row, end_row=end_row,
                               start_column=start_col, end_column=end_col)

            # 应用边框和居中对齐
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    cell.border = thin_border
                    if podid_col and cell.column == podid_col:
                        cell.alignment = center_align
                    elif group_col and cell.column == group_col:
                        cell.alignment = center_align
                    elif device_col and cell.column == device_col:
                        cell.alignment = center_align
                    elif cell.column == 11:  # 第11列是描述列
                        cell.alignment = left_align
                    else:
                        cell.alignment = center_align

            # 设置交替行颜色
            for row_idx in range(2, ws.max_row + 1):
                if row_idx % 2 == 0:
                    fill_color = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                else:
                    fill_color = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

                for col_idx in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.fill = fill_color

        wb.save(filename)
        print(f"已应用Excel格式美化: {filename}")

    def export_all_connections(self, filename):
        """导出所有连接关系到单个Excel文件"""
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            try:
                summary_df = self.generate_summary_data()
                summary_df.to_excel(writer, sheet_name='网络设计摘要', index=False)

                server_view = self.generate_server_view()
                server_view.to_excel(writer, sheet_name='服务器连接表', index=False)

                switch_views = self.generate_switch_view()
                switch_views['参数网络'].to_excel(writer, sheet_name='参数网络连接表', index=False)
                switch_views['存储网络'].to_excel(writer, sheet_name='存储网络连接表', index=False)

            except Exception as e:
                print(f"生成连接表时出错: {e}")
                if len(writer.book.sheetnames) == 0:
                    writer.book.create_sheet("错误信息")
                raise

        self.apply_excel_formatting(filename)
        print(f"所有连接表已导出到: {filename}")

    def validate_topology(self):
        """自检拓扑合法性：检查服务器覆盖率、端口溢出、连接完整性"""
        print("\n" + "=" * 60)
        print("拓扑自检")
        print("=" * 60)

        errors = []

        # 1. 检查服务器覆盖率
        param_nic_total = self.num_servers * self.param_ports_per_server
        storage_nic_total = self.num_servers * self.storage_ports_per_server
        param_conn_count = 0
        storage_conn_count = 0
        servers_with_param = set()
        servers_with_storage = set()

        for server in self.servers:
            for conn in server.connections:
                if conn.a_device == server.name:
                    if "参数" in conn.a_port:
                        param_conn_count += 1
                        servers_with_param.add(server.name)
                    elif "存储" in conn.a_port:
                        storage_conn_count += 1
                        servers_with_storage.add(server.name)

        if param_conn_count != param_nic_total:
            errors.append(f"参数网连接数不足: {param_conn_count}/{param_nic_total}")
        if storage_conn_count != storage_nic_total:
            errors.append(f"存储网连接数不足: {storage_conn_count}/{storage_nic_total}")
        if len(servers_with_param) != self.num_servers:
            errors.append(f"参数网覆盖服务器数不足: {len(servers_with_param)}/{self.num_servers}")
        if len(servers_with_storage) != self.num_servers:
            errors.append(f"存储网覆盖服务器数不足: {len(servers_with_storage)}/{self.num_servers}")

        # 2. 检查交换机端口是否溢出
        all_switches = (self.param_leaves + self.param_spines + self.param_cores +
                       self.storage_leaves + self.storage_spines + self.storage_cores)
        for sw in all_switches:
            max_ports = sw.max_ports
            used_ports = len(sw.connections)
            if used_ports > max_ports:
                errors.append(f"{sw.name} 端口溢出: {used_ports}/{max_ports}")

        # 3. 验证收敛比
        if self.param_3tier_needed:
            if self.param_spine_count != self.param_leaf_count:
                errors.append(f"参数网Leaf:Spine比例不为1:1 ({self.param_leaf_count}:{self.param_spine_count})")
        if self.storage_3tier_needed:
            if self.storage_spine_count != self.storage_leaf_count:
                errors.append(f"存储网Leaf:Spine比例不为1:1 ({self.storage_leaf_count}:{self.storage_spine_count})")

        if errors:
            print("发现以下问题:")
            for e in errors:
                print(f"  [错误] {e}")
        else:
            print("拓扑自检通过 - 无端口溢出，服务器全覆盖")

        return len(errors) == 0

    def print_summary(self):
        """打印网络设计摘要"""
        print("\n" + "=" * 60)
        print("网络设计摘要")
        print("=" * 60)
        print(f"GPU服务器数量: {self.num_servers}")
        print(f"每台服务器参数网卡数: {self.param_ports_per_server}")
        print(f"每台服务器存储网卡数: {self.storage_ports_per_server}\n")

        # 计算参数网络端口使用率（下联/上联各占switch_ports//2）
        half_ports = self.param_switch_ports // 2
        if self.param_3tier_needed:
            servers_per_group = min(self.param_servers_per_pod // self.param_ports_per_server, half_ports)
            param_downlink_usage = f"{servers_per_group}/{half_ports} ({servers_per_group / half_ports * 100:.1f}%)"
            param_uplink_usage = f"{half_ports}/{half_ports} (100.0%)"
            param_port_usage = f"下联:{param_downlink_usage}, 上联:{param_uplink_usage}"
        else:
            param_downlink_usage = f"{self.param_servers_per_group}/{half_ports} ({self.param_servers_per_group / half_ports * 100:.1f}%)"
            uplinks_used = min(half_ports, self.param_spine_count)
            param_uplink_usage = f"{uplinks_used}/{half_ports} ({uplinks_used / half_ports * 100:.1f}%)"
            param_port_usage = f"下联:{param_downlink_usage}, 上联:{param_uplink_usage}"

        # 计算存储网络端口使用率
        half_ports = self.storage_switch_ports // 2
        if self.storage_3tier_needed:
            servers_per_group = min(self.storage_servers_per_pod // self.storage_ports_per_server, half_ports)
            storage_downlink_usage = f"{servers_per_group}/{half_ports} ({servers_per_group / half_ports * 100:.1f}%)"
            storage_uplink_usage = f"{half_ports}/{half_ports} (100.0%)"
            storage_port_usage = f"下联:{storage_downlink_usage}, 上联:{storage_uplink_usage}"
        else:
            storage_downlink_usage = f"{self.storage_servers_per_group}/{half_ports} ({self.storage_servers_per_group / half_ports * 100:.1f}%)"
            uplinks_used = min(half_ports, self.storage_spine_count)
            storage_uplink_usage = f"{uplinks_used}/{half_ports} ({uplinks_used / half_ports * 100:.1f}%)"
            storage_port_usage = f"下联:{storage_downlink_usage}, 上联:{storage_uplink_usage}"

        # 打印参数网络信息
        print("参数网络:")
        print(f"  Leaf交换机数量: {self.param_leaf_count}")
        print(f"  Spine交换机数量: {self.param_spine_count}")
        if self.param_3tier_needed:
            print(f"  Core交换机数量: {self.param_core_count}")
            print(f"  网络层级: 3层(Leaf-Spine-Core)")
            print(f"  POD数量: {self.param_pods}, 每POD服务器数: {self.param_servers_per_pod}")
            print(f"  收敛比例: 1:1:1")
        else:
            print(f"  网络层级: 2层(Leaf-Spine)")
            print(f"  组数量: {self.param_groups}, 每组服务器数: {self.param_servers_per_group}")
            print(f"  收敛比例: 1:1")
        print(f"  交换机端口数: {self.param_switch_ports}")
        print(f"  端口使用率: {param_port_usage}")
        print(f"  网络速度: {self.param_speed}\n")

        # 打印存储网络信息
        print("存储网络:")
        print(f"  Leaf交换机数量: {self.storage_leaf_count}")
        print(f"  Spine交换机数量: {self.storage_spine_count}")
        if self.storage_3tier_needed:
            print(f"  Core交换机数量: {self.storage_core_count}")
            print(f"  网络层级: 3层(Leaf-Spine-Core)")
            print(f"  POD数量: {self.storage_pods}, 每POD服务器数: {self.storage_servers_per_pod}")
            print(f"  收敛比例: 1:1:1")
        else:
            print(f"  网络层级: 2层(Leaf-Spine)")
            print(f"  组数量: {self.storage_groups}, 每组服务器数: {self.storage_servers_per_group}")
            print(f"  收敛比例: 1:1")
        print(f"  交换机端口数: {self.storage_switch_ports}")
        print(f"  端口使用率: {storage_port_usage}")
        print(f"  网络速度: {self.storage_speed}")
        print("=" * 60)


# 示例使用
if __name__ == "__main__":
    # 使用配置文件初始化
    designer = NetworkDesigner("network_config.ini")
    designer.print_summary()
    designer.validate_topology()

    # 导出连接表
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M")
    output_filename = f"AI智算网络连接表_{timestamp}.xlsx"
    designer.export_all_connections(output_filename)